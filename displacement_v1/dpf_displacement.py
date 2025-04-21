from ansys.dpf import core as dpf
import numpy as np
import os
import openpyxl
import configparser
import time
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment

properties = configparser.ConfigParser()  # 클래스 객체 생성


def isdir_and_make(dir_name: str):
    if not (os.path.isdir(dir_name)):
        os.makedirs(name=dir_name, exist_ok=True)
        print(f"Success: Create {dir_name}\n")


def process_pairs_and_differences(arr: np.array):
    # 원본 배열의 길이가 짝수인지 확인
    if len(arr) % 2 != 0:
        # 홀수인 경우, 마지막 요소를 제외
        arr = arr[:-1]

    # 1. 두 개씩 데이터 추출 (2열로 재구성)
    pairs = arr.reshape(-1, 2)

    # 2. 각 쌍의 차이 계산 A-B
    differences = pairs[:, 0] - pairs[:, 1]

    # 3. 원본 쌍과 차이를 합치기
    result = list(np.column_stack((pairs, differences)).flatten())

    return result


def plot_columns_in_groups(df, filepath,  columns_per_plot=3):
    """
    DataFrame의 컬럼들을 지정된 수만큼 그룹으로 나누어 그래프를 그립니다.

    Parameters:
    -----------
    df : pandas DataFrame
        그래프를 그릴 데이터프레임
    columns_per_plot : int
        각 그래프당 표시할 컬럼 수
    figsize : tuple
        각 그래프의 크기 (가로, 세로)
    """

    # 모든 컬럼 가져오기
    all_columns = df.columns.tolist()

    # 컬럼들을 지정된 수만큼의 그룹으로 나누기
    number_of_groups = len(all_columns) // columns_per_plot

    # 색상 설정
    colors = ['red', 'blue', 'green']  # A, B, A-B

    # 각 그룹별로 그래프 그리기
    for i in range(number_of_groups):
        start_idx = i * columns_per_plot
        end_idx = start_idx + columns_per_plot
        current_columns = all_columns[start_idx:end_idx]
        title_name = current_columns[0].replace('-A', '')
        plt.figure(figsize=(12, 6))

        # 각 컬럼에 대해 라인 플롯 생성
        for j, column in enumerate(current_columns):
            plt.plot(df.index, df[column],
                     label=column,
                     color=colors[j],
                     linewidth=2,
                     marker='o',
                     markersize=4)

        # 그래프 꾸미기
        plt.title(f"{title_name} result", fontsize=18, pad=15)
        plt.xlabel('Time(sec)', fontsize=12)
        plt.ylabel('Displacement(mm)', fontsize=12)
        plt.legend(fontsize=10, bbox_to_anchor=(1.01, 1), loc='upper left')
        plt.grid(True, linestyle='--', alpha=0.7)

        # 여백 조정
        plt.tight_layout()
        plt.savefig(f'{filepath}/{title_name}.png', format='png')
        # plt.savefig(f'{filepath}/{filename}.svg', format='svg')
        plt.cla()  # clear the current axes
        plt.clf()  # clear the current figure
        plt.close()  # closes the current figure


def merge_repeated_headers(excel_path: str, sheet_names: list):
    """
    엑셀 파일의 첫 번째 행에서 반복되는 값들을 병합합니다.

    Parameters:
    -----------
    excel_path : str
        입력 엑셀 파일 경로
    output_path : str
        저장할 엑셀 파일 경로
    """
    # 엑셀 파일을 openpyxl로 읽기
    wb = openpyxl.load_workbook(excel_path)

    for sh in sheet_names:
        ws = wb[sh]

        # 첫 번째 행의 값들 가져오기
        first_row = list(ws[1])

        # 병합할 셀 범위 찾기
        merge_ranges = []
        start_col = 0
        current_value = first_row[0].value

        for i in range(1, len(first_row)):
            if first_row[i].value != current_value:
                if i - start_col > 1:  # 2개 이상의 셀이 같은 값을 가질 때만 병합
                    merge_ranges.append((start_col + 1, i))
                start_col = i
                current_value = first_row[i].value

        # 마지막 범위 처리
        if len(first_row) - start_col > 1:
            merge_ranges.append((start_col + 1, len(first_row)))

        # 셀 병합 및 스타일 적용
        for start, end in merge_ranges:
            merge_range = f"{chr(64 + start)}1:{chr(64 + end)}1"
            ws.merge_cells(merge_range)
            merged_cell = ws[f"{chr(64 + start)}1"]
            merged_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A1:A2')  # 첫번째 열에서의 Point 병합
        # 병합된 셀에 정렬 설정
        merged_cell = ws['A1']
        merged_cell.alignment = Alignment(horizontal='center',
                                          vertical='center')
    # 결과 저장
    wb.save(excel_path)


class DPF_Displacement:
    def __init__(self, result_path, input_file, rst_path, save_data, save_figure):
        isdir_and_make(result_path)  # result 폴더 생성

        self.model = None
        self.mesh = None
        self.result_path = result_path
        self.input_file = input_file
        self.rst_path = rst_path
        self.save_data = True if save_data == 'Yes' else False
        self.save_figure = True if save_figure == 'Yes' else False

        self.exe_time = None
        self.sheet_name = []
        self.points = {}
        self.nodes = {}
        self.time_sec = []
        self.step_ids = []
        self.coordinates = {}

    def run(self):
        print("************************************************************")
        print("*** DPF Displacement Analysis Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        print(f"[INFO] Save CSV data option: {self.save_data}")
        print(f"[INFO] Save Figure option: {self.save_figure}")
        print(f"[INFO] Result Location: {self.result_path}")
        print(f"[INFO] RST folder: {self.rst_path}\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')

        self._load_in_point_node()
        self._load_rst()
        self._get_time_frequencies()

        if self.save_data is True or self.save_figure is True:
            isdir_and_make(f'{self.result_path}/{self.exe_time}')  # result 결과 시간 폴더 생성
            self._save_coordinates()  # 좌표 csv 저장

        dict_dis = self._extract_step_displacement()

        # 1. 파일 생성
        writer = pd.ExcelWriter(f"{self.result_path}/result_disp_{self.exe_time}.xlsx", engine='openpyxl')
        for sn in self.sheet_name:
            df_coord = self._get_dataframe_coordinates(sn)
            df_coord.to_excel(writer, sheet_name=f'coordinate_{sn}', index=False, header=False)

        displacement_axis = ['X', 'Y']
        diff_sheet = []
        for sn in self.sheet_name:
            for axis in displacement_axis:
                df_disp, diff_state = self.get_dataframe_dis(dict_dis, axis, sn)
                disp_sheet_name = f"disp_{axis}_{sn}"
                df_disp.to_excel(writer, sheet_name=disp_sheet_name, index=False, header=False)
                if diff_state is True:
                    diff_sheet.append(disp_sheet_name)

        # 3. 작성 완료 후 파일 저장
        writer.close()

        # 4. 일부 쉬트 merge 셀 적용
        if diff_sheet:
            merge_repeated_headers(f"{self.result_path}/result_disp_{self.exe_time}.xlsx", diff_sheet)

        self.stop()

    def stop(self):
        print(f'Ending at: {time.strftime("%a, %d-%b-%Y %I:%M:%S", time.localtime(time.time()))}')
        print("************************************************************")
        print("*** DPF Displacement Analysis completed")
        print("************************************************************\n")

    def get_dataframe_dis(self, dict_dis: dict, axis: str, sn: str):
        time_data = dict_dis.keys()
        disp = [[df.loc[node, axis] for node in self.nodes[sn]] for df in dict_dis.values()]
        if '-A' in self.points[sn][0] or '-B' in self.points[sn][0]:
            diff_apply = True
            node_by_point = dict(zip(self.points[sn], self.nodes[sn]))
            organized, organized_diff, organized_nodes, point_main, point_sub, units = [], [], [], [], [], []
            for point in self.points[sn]:
                if '-A' in point:
                    B_point = point.replace('-A', '-B')
                    organized.append(point)
                    organized.append(B_point)

                    organized_diff.append(point)
                    organized_diff.append(B_point)
                    organized_diff.append(point.replace('-A', 'A-B'))

                    organized_nodes.append(node_by_point[point])
                    organized_nodes.append(node_by_point[B_point])
                    organized_nodes.append('')

                    point_sub.append('A')
                    point_sub.append('B')
                    point_sub.append('A-B')

                    units.append(f'Disp[{self.mesh.unit}]')
                    units.append(f'Disp[{self.mesh.unit}]')
                    units.append(f'Diff.[{self.mesh.unit}]')

                    name = point.replace('-A', '')
                    for _ in range(3):
                        point_main.append(name)

            lst_disp_diff = [process_pairs_and_differences(values) for values in pd.DataFrame(data=disp, columns=self.points[sn]).reindex(columns=organized).values]
            # points간 diff 포함 그래프 그리기
            if self.save_figure is True:
                plot_columns_in_groups(df=pd.DataFrame(data=lst_disp_diff, columns=organized_diff, index=list(time_data)), filepath=f'{self.result_path}/{self.exe_time}')

            point_col1 = ['Points'] + point_main
            point_col2 = ['Points'] + point_sub
            node_col = ['Node'] + organized_nodes
            name_col = ['Time'] + units

            disp_rev = [point_col1, point_col2, node_col, name_col]
            for t, dis in zip(time_data, lst_disp_diff):
                disp_rev.append([t] + dis)
        else:
            diff_apply = False
            point_col = ['Points'] + self.points[sn]
            node_col = ['Node'] + self.nodes[sn]
            name_col = ['Time'] + [f'Disp[{self.mesh.unit}]' for _ in range(len(self.nodes[sn]))]

            disp_rev = [point_col, node_col, name_col]
            for t, dis in zip(time_data, disp):
                disp_rev.append([t] + dis)

        return pd.DataFrame(disp_rev), diff_apply

    def _load_in_point_node(self):
        wb = openpyxl.load_workbook(self.input_file)
        self.sheet_name = wb.sheetnames  # the list that stores the sheetnames
        for sn in self.sheet_name:
            df_input = pd.read_excel(self.input_file, engine='openpyxl', sheet_name=sn)
            self.points[sn] = list(df_input.iloc[:, 0])
            self.nodes[sn] = list(df_input.iloc[:, 1])

    def _load_rst(self):
        # rst 파일 찾기
        rst_file = [file for file in os.listdir(self.rst_path) if '.rst' in file]
        # rst 파일 직접 로드
        ds = dpf.DataSources(os.path.join(self.rst_path, rst_file[0]))
        self.model = dpf.Model(ds)
        print(self.model)

    def _save_coordinates(self):
        df_node_coord = pd.DataFrame(data=[[node.id] + node.coordinates for node in self.mesh.nodes], columns=['Node ID', 'Location X', 'Location Y', 'Location Z'])
        print(df_node_coord)
        df_node_coord.to_csv(f'{self.result_path}/{self.exe_time}/node_coordinate.csv', index=False)

    def _get_time_frequencies(self):
        metadata = self.model.metadata
        self.mesh = metadata.meshed_region

        self.coordinates = {node.id: node.coordinates for node in self.mesh.nodes}

        # time_frequencies 객체 가져오기
        time_freq = metadata.time_freq_support.time_frequencies
        self.time_sec = time_freq.data
        self.step_ids = time_freq.scoping.ids

    def _get_dataframe_coordinates(self, sn):
        data_coord = [['Points', 'Node', 'Location X', 'Location Y']]
        for point, node in zip(self.points[sn], self.nodes[sn]):
            coordinates = self.coordinates[node]
            data_coord.append([point, node, coordinates[0], coordinates[1]])
        return pd.DataFrame(data_coord)

    def _extract_step_displacement(self):
        disp_op = self.model.results.displacement()
        time_index = 0
        dict_dis = {}
        for step in self.step_ids:
            time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)
            # disp = model.results.displacement.on_time_scoping(time_steps_scoping).eval()
            disp_op.inputs.time_scoping(time_steps_scoping)
            disp = disp_op.outputs.fields_container()
            for dis in disp:
                if float(self.time_sec[time_index]).is_integer():
                    node_ids = dis.scoping.ids
                    df = pd.DataFrame(data=dis.data, index=node_ids, columns=['X', 'Y', 'Z'])
                    df_dis = df.sort_index(ascending=True)
                    if self.save_data is True:
                        df_dis.to_csv(f'{self.result_path}/{self.exe_time}/displacement_step{step}_{round(self.time_sec[time_index], 1)}s.csv')
                    dict_dis[int(self.time_sec[time_index])] = df_dis
                time_index += 1
        return dict_dis


if __name__ == "__main__":
    properties.read('config.ini', encoding="UTF-8")  # 파일 읽기
    fm_params = properties["ENV"]  # 섹션 선택

    dpf_analysis = DPF_Displacement(result_path=fm_params['result_path'].strip(),
                                    input_file=fm_params['input_file_path'].strip(),
                                    rst_path=fm_params['rst_folder_path'].strip(),
                                    save_data=fm_params['save_raw_data'].strip(),
                                    save_figure=fm_params['save_figure'].strip())
    dpf_analysis.run()
