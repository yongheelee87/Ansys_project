import numpy as np
from ansys.dpf import core as dpf
from ansys.dpf.core import operators as ops
import os
import openpyxl
import time
import yaml
import pandas as pd
from commons import isdir_and_make, add_ys_uts, get_max_value, concat_dataframes_with_gaps, read_excel_with_merged_cells, find_changed_points, rainflow_point, rainflow_counting


ROOT_PATH = os.getcwd()  # 현재 위치 가져오기
RESULT_PATH = f"{ROOT_PATH}/result"
DATA_PATH = f"{ROOT_PATH}/data"
REFERENCE_EXCEL = f"{DATA_PATH}/Calculation_Reference.xlsx"
RESULT_EXCEL = f"{DATA_PATH}/Result_Format.xlsx"
ABBREVIATION = f"{DATA_PATH}/abbreviation.yaml"
NAME_SELECTION_PREFIX = "S_"
COLUMNS = ["Time", "NodeID", "Stress", "Temperature"]
COLUMNS_SIGN = ["Time", "NodeID", "Max_Stress", "Temperature", "Principal1"]
RESULT_COLUMNS = ["Part", "Time", "Max_Temp", "NodeID_T", "Stress", "NodeID_S", "YS", "UTS", "Temperature"]
CASES = ['Max EQ of Each Step', 'Max Temp of Each Step', 'Max EQ', 'Max Temp', 'Max Principal', 'Min Principal', 'Signed Max EQ']
RAINFLOW_ONOFF = True


class DPF_Stress_Temperature:
    def __init__(self, save_data, rst_path):
        isdir_and_make(RESULT_PATH)  # result 폴더 생성
        self.save_data = save_data
        self.abbre_mat = None
        self.model = None
        self.stress_op = None
        self.temperature_op = None
        self.mesh = None
        self.metadata = None
        self.bodies = []
        self.result_part = {}
        self.result_time = []
        self.rst_path = rst_path
        self.sheet_name = []
        self.time_sec = []
        self.step_ids = []
        # self.coordinates = {}
        self.materials = {}
        
        # 초기화 함수
        self._load_abbreviation_material()  # 약자 읽기
        self._load_rst()  # model, rst 읽기

    def reload_model(self, input_rst_path):
        self.rst_path = input_rst_path
        self._load_rst()  # model, rst 읽기
        print(f"설정된 경로: {self.rst_path}\n")

    def analyze_base_model(self):
        print("************************************************************")
        print("*** #1.Mode DPF Stress Temperature Analysis Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {exe_time}')
        if self.save_data is True:
            isdir_and_make(f'{RESULT_PATH}/{exe_time}')  # result 결과 시간 폴더 생성
        isdir_and_make(f'{RESULT_PATH}/Mode 1')  # result 결과 시간 폴더 생성

        self._load_result_format()
        self._load_part()

        self.set_operation()
        df_sum = self.wrap_result(exe_time)
        self.write_result_xlsx(df_sum, exe_time)

        self.stop('#1.Mode')
        
    def check_temperature_range(self):
        print("************************************************************")
        print("*** #2.Mode Find Temperature Range Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {exe_time}')
        isdir_and_make(f'{RESULT_PATH}/Mode 2')  # result 폴더 생성

        self._load_part()
        self.set_operation()
        self.wrap_temperature_result(exe_time)

        self.stop('#2.Mode')

    def analyze_base_model_all_nodes(self):
        print("************************************************************")
        print("*** #3.Mode DPF Stress Temperature Analysis with All nodes Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {exe_time}')

        isdir_and_make(f'{RESULT_PATH}/Mode 3')  # result 폴더 생성

        self._load_part()
        self.set_operation()
        self.wrap_all_nodes_result(exe_time)

        self.stop('#3.Mode')

    def analyze_base_model_specific_nodes(self, material, nodes):
        print("************************************************************")
        print("*** #4.Mode DPF Stress Temperature Analysis with Specific nodes Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {exe_time}')

        isdir_and_make(f'{RESULT_PATH}/Mode 4')  # result 폴더 생성

        self.set_operation()
        self.wrap_specific_nodes_result(exe_time, material, nodes)

        self.stop('#4.Mode')


    def stop(self, mode: str):
        print(f'Ending at: {time.strftime("%a, %d-%b-%Y %I:%M:%S", time.localtime(time.time()))}')
        print("************************************************************")
        print(f"*** {mode} DPF Stress Temperature Analysis completed")
        print("************************************************************\n")

    def _load_part(self):
        # Body 가져오기
        self.bodies = [body_name for body_name in self.metadata.available_named_selections if NAME_SELECTION_PREFIX in body_name[:2]]
        print(f"[INFO] Selected Names: {self.bodies}\n")

        # self.coordinates = {node.id: node.coordinates for node in self.mesh.nodes}
        self.materials = {'_'.join(body.split('_')[2:]): body.split('_')[1] for body in self.bodies}

    def set_operation(self):
        # time_frequencies 객체 가져오기
        time_freq = self.metadata.time_freq_support.time_frequencies
        self.time_sec = time_freq.data
        self.step_ids = time_freq.scoping.ids

        self.stress_op = self.model.results.stress()
        self.stress_op.inputs.requested_location.connect('Nodal')

        self.temperature_op = self.model.results.structural_temperature()
        self.temperature_op.inputs.requested_location.connect('Nodal')

    def wrap_result(self, exe_time):
        isdir_and_make(f'{RESULT_PATH}/Mode 1/{exe_time}')

        # 1. 파일 생성
        writer = pd.ExcelWriter(f'{RESULT_PATH}/Mode 1/{exe_time}/result_stress_temp.xlsx', engine='openpyxl')
        if RAINFLOW_ONOFF:
            writer_rainflow = pd.ExcelWriter(f'{RESULT_PATH}/Mode 1/{exe_time}/result_rainflow_point.xlsx', engine='openpyxl')
            writer_signed_rainflow = pd.ExcelWriter(f'{RESULT_PATH}/Mode 1/{exe_time}/result_signed_rainflow_point.xlsx', engine='openpyxl')
            writer_cylcle = pd.ExcelWriter(f'{RESULT_PATH}/Mode 1/{exe_time}/result_cycle_rainflow_point.xlsx', engine='openpyxl')
            writer_cycle_signed = pd.ExcelWriter(f'{RESULT_PATH}/Mode 1/{exe_time}/result_cycle_signed_rainflow_point.xlsx', engine='openpyxl')

        # 2. 쉬트 작성
        summary_data = []
        for body in self.bodies:
            part_name = '_'.join(body.split('_')[2:])  # part name 추출 (추가 작업 필요)
            material = self.abbre_mat[body.split('_')[1]]
            lst_df_st_temp_rev = [add_ys_uts(material, REFERENCE_EXCEL, df_st_temp, case) for df_st_temp, case in zip(self._extract_all_stress_temperature(body, exe_time), CASES)]
            if RAINFLOW_ONOFF:
                df_max = lst_df_st_temp_rev[2].iloc[1:].reset_index(drop=True)
                changed_indices = find_changed_points(df_max['Stress'].astype(float))
                reversal_values = [df_max['Stress'].values[0]] + df_max.loc[changed_indices, 'Stress'].values.tolist() + [df_max['Stress'].values[-1]]
                df_cycle = rainflow_counting([float(val) for val in reversal_values])
                df_cycle.to_excel(writer_cylcle, sheet_name=part_name, index=False, header=True)
                df_max.loc[changed_indices].to_excel(writer_rainflow, sheet_name=part_name, index=False, header=True)
                rainflow_point(result_path=f'{RESULT_PATH}/Mode 1/{exe_time}', part_name=part_name, df=df_max, changed_indices=changed_indices)

                df_max = lst_df_st_temp_rev[-1].iloc[1:].reset_index(drop=True)
                changed_indices = find_changed_points(df_max['Stress'].astype(float))
                reversal_values = [df_max['Stress'].values[0]] + df_max.loc[changed_indices, 'Stress'].values.tolist() + [df_max['Stress'].values[-1]]
                df_cycle = rainflow_counting([float(val) for val in reversal_values])
                df_cycle.to_excel(writer_cycle_signed, sheet_name=part_name, index=False, header=True)
                df_max.loc[changed_indices].to_excel(writer_signed_rainflow, sheet_name=part_name, index=False, header=True)
                rainflow_point(result_path=f'{RESULT_PATH}/Mode 1/{exe_time}', part_name=f'{part_name}_Signed', df=df_max, changed_indices=changed_indices)

            for t in self.result_time:
                max_stress_by_time = lst_df_st_temp_rev[0][lst_df_st_temp_rev[0]['Time'] == t].values[0]
                max_temperature_by_time = lst_df_st_temp_rev[1][lst_df_st_temp_rev[1]['Time'] == t].values[0]
                '''
                # 파트명 시간 온도 위치 스트레스 위치 YS UTS 온도
                [part_name, t, max_temperature_by_time['Temperature'], max_temperature_by_time['NodeID'],
                 max_stress_by_time['Stress'], max_stress_by_time['NodeID'], max_stress_by_time['YS'], max_stress_by_time['UTS'], max_temperature_by_time['Temperature']]
                '''
                summary_data.append([part_name, t, max_temperature_by_time[4], max_temperature_by_time[2],
                                     max_stress_by_time[3], max_stress_by_time[2], max_stress_by_time[5], max_stress_by_time[6], max_stress_by_time[4]])
            df_st_temp = concat_dataframes_with_gaps(lst_df_st_temp_rev)  # 세로 형태로 가운데에 빈공간을 넣고 합침
            df_st_temp.to_excel(writer, sheet_name=part_name, index=False, header=False)

        # 3. 작성 완료 후 파일 저장
        writer.close()
        if RAINFLOW_ONOFF:
            writer_rainflow.close()
            writer_signed_rainflow.close()
            writer_cylcle.close()
            writer_cycle_signed.close()
        return pd.DataFrame(summary_data, columns=RESULT_COLUMNS)  # 결과 추출위한 데이터프레임

    def write_result_xlsx(self, df, exe_time):
        wb = openpyxl.load_workbook(RESULT_EXCEL)
        for sh, parts in self.result_part.items():
            ws = wb[sh]
            time_source = sh.split('_')
            df_time = df[df['Time'] == int(time_source[0].replace('s', ''))]
            if 'Temp' in time_source[-1]:
                for pos, part in parts.items():
                    if pos[1] != 1:  # 첫번째 열은 KTP 포함 제외
                        part_data = df_time[df_time['Part'] == part].values[0]
                        ws.cell(row=pos[0] + 1, column=pos[1]).value = part_data[2]  # Temperature
                        ws.cell(row=pos[0] + 2, column=pos[1]).value = part_data[3]  # Node ID
            else:
                for pos, part in parts.items():
                    if pos[1] != 1:  # 첫번째 열은 KTP 포함 제외
                        part_data = df_time[df_time['Part'] == part].values[0]
                        ws.cell(row=pos[0] + 1, column=pos[1]).value = self.abbre_mat[self.materials[part]]  # 소재
                        ws.cell(row=pos[0] + 2, column=pos[1]).value = part_data[6]  # YS
                        ws.cell(row=pos[0] + 3, column=pos[1]).value = part_data[7]  # UTS
                        ws.cell(row=pos[0] + 4, column=pos[1]).value = part_data[4]  # Stress
                        ws.cell(row=pos[0] + 5, column=pos[1]).value = part_data[8]  # Temperature
                        ws.cell(row=pos[0] + 6, column=pos[1]).value = part_data[5]  # Node ID

        wb.save(f'{RESULT_PATH}//Mode 1/{exe_time}/result_summary.xlsx')
        wb.close()

    def wrap_temperature_result(self, exe_time):
        # 1. 파일 생성
        writer = pd.ExcelWriter(f'{RESULT_PATH}/Mode 2/{exe_time}_result_temperature.xlsx', engine='openpyxl')

        lst_data = []
        for body in self.bodies:
            lst_temperature = []

            part_name = '_'.join(body.split('_')[2:])  # part name 추출 (추가 작업 필요)
            mesh_scoping = self.metadata.named_selection(body)
            self.temperature_op.inputs.mesh_scoping.connect(mesh_scoping)

            time_index = 0
            for step in self.step_ids:
                time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)

                # temperature_op 연결 및 필드 구성
                self.temperature_op.inputs.time_scoping(time_steps_scoping)
                temperature_field = self.temperature_op.outputs.fields_container()
                for temperature in temperature_field:
                    print(temperature)
                    if float(self.time_sec[time_index]).is_integer():
                        lst_temperature.append(temperature.data)
                    time_index += 1
            arr_temperature = np.array(lst_temperature).flatten()
            lst_data.append([part_name, np.min(arr_temperature), np.max(arr_temperature)])
        df_temperature = pd.DataFrame(lst_data, columns=['파트명', '최소온도', '최대온도'])
        df_temperature.to_excel(writer, sheet_name='온도', index=False, header=True)

        # 2. 작성 완료 후 파일 저장
        writer.close()

    def wrap_all_nodes_result(self, exe_time):
        isdir_and_make(f'{RESULT_PATH}/Mode 3/{exe_time}')  # result 폴더 생성

        # 1. 파일 생성
        writer_sum = pd.ExcelWriter(f'{RESULT_PATH}/Mode 3/{exe_time}_result_stress_temp_all_nodes_summary.xlsx', engine='openpyxl')

        # 2. 쉬트 작성
        for body in self.bodies:
            lst_df_st_temp, lst_check_nodes, lst_fail_nodes, lst_time = [], [], [], []
            part_name = '_'.join(body.split('_')[2:])  # part name 추출 (추가 작업 필요)
            # 파트 파일 생성
            writer = pd.ExcelWriter(f'{RESULT_PATH}/Mode 3/{exe_time}/{part_name}.xlsx', engine='openpyxl')
            material = self.abbre_mat[body.split('_')[1]]

            mesh_scoping = self.metadata.named_selection(body)
            self.stress_op.inputs.mesh_scoping.connect(mesh_scoping)
            self.temperature_op.inputs.mesh_scoping.connect(mesh_scoping)

            time_index = 0
            for step in self.step_ids:
                time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)

                # Equivalent Stress 연결 및 필드 구성
                self.stress_op.inputs.time_scoping(time_steps_scoping)
                eqv_op = ops.invariant.von_mises_eqv_fc(self.stress_op)
                eqv_field = eqv_op.outputs.fields_container()

                # temperature_op 연결 및 필드 구성
                self.temperature_op.inputs.time_scoping(time_steps_scoping)
                temperature_field = self.temperature_op.outputs.fields_container()

                for eqv_stress, temperature in zip(eqv_field, temperature_field):
                    if float(self.time_sec[time_index]).is_integer():
                        stress_node_ids = eqv_stress.scoping.ids
                        temperature_node_ids = temperature.scoping.ids

                        df_stress = pd.DataFrame(data=eqv_stress.data, index=stress_node_ids, columns=['Stress']).sort_index(ascending=True)
                        df_temperature = pd.DataFrame(data=temperature.data, index=temperature_node_ids, columns=['Temperature']).sort_index(ascending=True)
                        df_st_temp = pd.concat([df_stress, df_temperature], axis=1)

                        time_sec = round(self.time_sec[time_index], 1)
                        df_st_temp['Time'] = time_sec
                        df_st_temp['NodeID'] = df_st_temp.index.values
                        df_st_temp = df_st_temp[COLUMNS]
                        df_st_temp = add_ys_uts(material, REFERENCE_EXCEL, df_st_temp, case='')

                        lst_time.append(time_sec)
                        check_nodes = list(df_st_temp['NodeID'][df_st_temp['Note'] == 'Check'].values)
                        fail_nodes = list(df_st_temp['NodeID'][df_st_temp['Note'] == 'Fail'].values)
                        lst_check_nodes.append(str(check_nodes).replace('[', '').replace(']', '') if len(check_nodes) != 0 else '')
                        lst_fail_nodes.append(str(fail_nodes).replace('[', '').replace(']', '') if len(fail_nodes) != 0 else '')

                        df_st_temp.to_excel(writer, sheet_name=f'{int(time_sec)}s', index=False, header=True)
                    time_index += 1

            df_summary = pd.DataFrame(lst_time, columns=['Time'])
            df_summary['Check_Nodes'] = lst_check_nodes
            df_summary['Fail_Nodes'] = lst_fail_nodes
            df_summary.to_excel(writer_sum, sheet_name=part_name, index=False, header=True)

            writer.close()  # 작성 완료 후 파일 저장

        # 3. 작성 완료 후 파일 저장
        writer_sum.close()

    def wrap_specific_nodes_result(self, exe_time: str, mat: str, node_ids: str):
        # 1. 파일 생성
        writer = pd.ExcelWriter(f'{RESULT_PATH}/Mode 4/{exe_time}_result_stress_temp_specific_nodes.xlsx', engine='openpyxl')

        lst_df_st_temp, lst_check_nodes, lst_fail_nodes, lst_time = [], [], [], []

        material = self.abbre_mat[mat.strip()]
        lst_node = [int(node.strip()) for node in node_ids.split(',')]
        mesh_scoping = dpf.mesh_scoping_factory.nodal_scoping(lst_node)

        self.stress_op.inputs.mesh_scoping.connect(mesh_scoping)
        self.temperature_op.inputs.mesh_scoping.connect(mesh_scoping)

        time_index = 0
        for step in self.step_ids:
            time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)

            # Equivalent Stress 연결 및 필드 구성
            self.stress_op.inputs.time_scoping(time_steps_scoping)
            eqv_op = ops.invariant.von_mises_eqv_fc(self.stress_op)
            eqv_field = eqv_op.outputs.fields_container()

            # temperature_op 연결 및 필드 구성
            self.temperature_op.inputs.time_scoping(time_steps_scoping)
            temperature_field = self.temperature_op.outputs.fields_container()

            for eqv_stress, temperature in zip(eqv_field, temperature_field):
                if float(self.time_sec[time_index]).is_integer():
                    stress_node_ids = eqv_stress.scoping.ids
                    temperature_node_ids = temperature.scoping.ids

                    df_stress = pd.DataFrame(data=eqv_stress.data, index=stress_node_ids, columns=['Stress']).sort_index(ascending=True)
                    df_temperature = pd.DataFrame(data=temperature.data, index=temperature_node_ids, columns=['Temperature']).sort_index(ascending=True)
                    df_st_temp = pd.concat([df_stress, df_temperature], axis=1)

                    time_sec = round(self.time_sec[time_index], 1)
                    df_st_temp['Time'] = time_sec
                    df_st_temp['NodeID'] = df_st_temp.index.values
                    df_st_temp = df_st_temp[COLUMNS]
                    df_st_temp = add_ys_uts(material, REFERENCE_EXCEL, df_st_temp, case='')

                    lst_time.append(time_sec)
                    check_nodes = list(df_st_temp['NodeID'][df_st_temp['Note'] == 'Check'].values)
                    fail_nodes = list(df_st_temp['NodeID'][df_st_temp['Note'] == 'Fail'].values)
                    lst_check_nodes.append(str(check_nodes).replace('[', '').replace(']', '') if len(check_nodes) != 0 else '')
                    lst_fail_nodes.append(str(fail_nodes).replace('[', '').replace(']', '') if len(fail_nodes) != 0 else '')
                    lst_df_st_temp.append(df_st_temp)
                time_index += 1

        df_total_st_temp = pd.concat(lst_df_st_temp, axis=0)  # 전부 합침
        df_part_st_temp = concat_dataframes_with_gaps(lst_df_st_temp)  # 세로 형태로 가운데에 빈공간을 넣고 합침
        df_part_st_temp.to_excel(writer, sheet_name='data', index=False, header=False)

        df_summary = pd.DataFrame(lst_time, columns=['Time'])
        df_summary['Check_Nodes'] = lst_check_nodes
        df_summary['Fail_Nodes'] = lst_fail_nodes
        df_summary.to_excel(writer, sheet_name='sum', index=False, header=True)

        for node in lst_node:
            df_total_st_temp[df_total_st_temp['NodeID'] == node].to_excel(writer, sheet_name=f'{node}', index=False, header=True)
        # 3. 작성 완료 후 파일 저장
        writer.close()

    def out_all_nodes_raw(self):
        self._load_part()
        self.set_operation()

        isdir_and_make(f'{RESULT_PATH}/Mode 5')  # result 폴더 생성

        # 1. 네임별 mesh
        for body in self.bodies:
            mesh_scoping = self.metadata.named_selection(body)
            print(mesh_scoping)

        time_index = 0
        for step in self.step_ids:
            time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)

            # Equivalent Stress 연결 및 필드 구성
            self.stress_op.inputs.time_scoping(time_steps_scoping)
            eqv_op = ops.invariant.von_mises_eqv_fc(self.stress_op)
            eqv_field = eqv_op.outputs.fields_container()

            # temperature_op 연결 및 필드 구성
            self.temperature_op.inputs.time_scoping(time_steps_scoping)
            temperature_field = self.temperature_op.outputs.fields_container()

            for eqv_stress, temperature in zip(eqv_field, temperature_field):
                if float(self.time_sec[time_index]).is_integer():
                    stress_node_ids = list(eqv_stress.scoping.ids)
                    temperature_node_ids = list(temperature.scoping.ids)

                    print(stress_node_ids)
                    print(temperature_node_ids)
                time_index += 1

    def  _extract_all_stress_temperature(self, selection, exe_time):
        mesh_scoping = self.metadata.named_selection(selection)
        self.stress_op.inputs.mesh_scoping.connect(mesh_scoping)
        self.temperature_op.inputs.mesh_scoping.connect(mesh_scoping)

        time_index = 0
        dict_st_temp, lst_stress, lst_temperature, tuple_stress, tuple_temperature, tuple_principal1, tuple_principal3 = {}, [], [], [], [], [], []
        for step in self.step_ids:
            time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)

            # Equivalent Stress 연결 및 필드 구성
            self.stress_op.inputs.time_scoping(time_steps_scoping)
            # Equivalent stress operator
            # self.eqv_op = stress_op.eqv()
            eqv_op = ops.invariant.von_mises_eqv_fc(self.stress_op)
            eqv_field = eqv_op.outputs.fields_container()

            # Principal Stress 연결 및 필드 구성
            principal_op = ops.invariant.principal_invariants_fc(self.stress_op)
            principal_stress_1 = principal_op.outputs.fields_eig_1()
            # principal_stress_2 = principal_op.outputs.fields_eig_2()
            principal_stress_3 = principal_op.outputs.fields_eig_3()

            # temperature_op 연결 및 필드 구성
            self.temperature_op.inputs.time_scoping(time_steps_scoping)
            temperature_field = self.temperature_op.outputs.fields_container()

            for eqv_stress, temperature, p_stress_1, p_stress_3 in zip(eqv_field, temperature_field, principal_stress_1, principal_stress_3):
                if float(self.time_sec[time_index]).is_integer():
                    stress_node_ids = eqv_stress.scoping.ids
                    temperature_node_ids = temperature.scoping.ids
                    p_stress_1_node_ids = p_stress_1.scoping.ids
                    p_stress_3_node_ids = p_stress_3.scoping.ids

                    df_stress = pd.DataFrame(data=eqv_stress.data, index=stress_node_ids, columns=['Equivalent_Stress']).sort_index(ascending=True)
                    df_temperature = pd.DataFrame(data=temperature.data, index=temperature_node_ids, columns=['Temperature']).sort_index(ascending=True)
                    df_p_stress_1 = pd.DataFrame(data=p_stress_1.data, index=p_stress_1_node_ids, columns=['Principal_1']).sort_index(ascending=True)
                    df_p_stress_3 = pd.DataFrame(data=p_stress_3.data, index=p_stress_3_node_ids, columns=['Principal_3']).sort_index(ascending=True)
                    df_st_temp = pd.concat([df_stress, df_temperature, df_p_stress_1, df_p_stress_3], axis=1)

                    time_sec = round(self.time_sec[time_index], 1)

                    lst_max, tuple_max = get_max_value(df_st_temp, time_sec)

                    lst_stress.append(lst_max[0][:4])  # 'Time', 'Node ID' 'Equivalent_Stress', 'Temperature'
                    lst_temperature.append(lst_max[1][:4])  # 'Time', 'Node ID' 'Equivalent_Stress', 'Temperature'
                    tuple_stress.append(tuple_max[0])
                    tuple_temperature.append(tuple_max[1])
                    tuple_principal1.append(tuple_max[2])
                    tuple_principal3.append(tuple_max[3])
                    if self.save_data is True:
                        df_st_temp.to_csv(f'{RESULT_PATH}/{exe_time}/{selection}_step{step}_{round(self.time_sec[time_index], 1)}s.csv')
                    dict_st_temp[int(self.time_sec[time_index])] = df_st_temp
                time_index += 1
        max_stress_node = max(tuple_stress, key=lambda x: x[1])[0]
        max_temperature_node = max(tuple_temperature, key=lambda x: x[1])[0]
        max_principal1_node = max(tuple_principal1, key=lambda x: x[1])[0]
        max_principal3_node = max(tuple_principal3, key=lambda x: x[1])[0]

        lst_max_stress, lst_max_temperature, lst_principal1, lst_principal3, lst_max_stress_sign = [], [], [], [], []
        for t, df_data in dict_st_temp.items():
            lst_max_stress.append([t, max_stress_node] + df_data.loc[max_stress_node, :].values.tolist()[:2])
            lst_max_temperature.append([t, max_temperature_node] + df_data.loc[max_temperature_node, :].values.tolist()[:2])
            data_principal1 = df_data.loc[max_principal1_node, :].values.tolist()
            data_principal3 = df_data.loc[max_principal3_node, :].values.tolist()
            lst_principal1.append([t, max_principal1_node, data_principal1[2], data_principal1[1]])
            lst_principal3.append([t, max_principal3_node, data_principal3[3], data_principal3[1]])
            lst_max_stress_sign.append([t, max_stress_node] + df_data.loc[max_stress_node, :].values.tolist()[:3])

        df_max_sign = pd.DataFrame(lst_max_stress_sign, columns=COLUMNS_SIGN)
        sign_values = np.sign(df_max_sign["Principal1"])
        sign_values = np.where(sign_values == 0, 1, sign_values)
        df_max_sign['Stress'] = sign_values * df_max_sign['Max_Stress']

        # df_stress, df_temperature, df_max_stress, df_max_temperature # max_principal # min_principal
        return [pd.DataFrame(lst_stress, columns=COLUMNS), pd.DataFrame(lst_temperature, columns=COLUMNS), pd.DataFrame(lst_max_stress, columns=COLUMNS),
                pd.DataFrame(lst_max_temperature, columns=COLUMNS), pd.DataFrame(lst_principal1, columns=COLUMNS), pd.DataFrame(lst_principal3, columns=COLUMNS), df_max_sign[COLUMNS]]

    def _load_abbreviation_material(self):
        with open(ABBREVIATION, encoding='utf-8-sig') as f:
            setting = yaml.load(f, Loader=yaml.SafeLoader)
        self.abbre_mat = setting['Material']

    def _load_rst(self):
        # rst 파일 찾기
        rst_file = [file for file in os.listdir(self.rst_path) if '.rst' in file]
        # rst 파일 직접 로드
        self.model = dpf.Model(os.path.join(self.rst_path, rst_file[0]))
        self.metadata = self.model.metadata
        self.mesh = self.metadata.meshed_region
        print(self.model)

    def _load_result_format(self):
        # 워크북 로드
        wb = openpyxl.load_workbook(RESULT_EXCEL)
        parts, list_time_in = {}, []
        for sheet_name in wb.sheetnames:
            list_time_in.append(int(sheet_name.split('_')[0].replace('s', '')))
            parts[sheet_name] = read_excel_with_merged_cells(wb[sheet_name])
        self.result_time = list(set(list_time_in))
        self.result_part = parts
