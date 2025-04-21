import os
import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from collections import deque


NOT_AVAILABLE_NUM = -999
KEYWORD = ['프로', '2D Layout']


def isdir_and_make(dir_name: str):
    if not (os.path.isdir(dir_name)):
        os.makedirs(name=dir_name, exist_ok=True)
        print(f"Success: Create {dir_name}\n")


def stress_conditions(row):
    if row['YS'] == NOT_AVAILABLE_NUM and row['UTS'] == NOT_AVAILABLE_NUM:
        val = 'N/A'
    elif row['YS'] == NOT_AVAILABLE_NUM or row['UTS'] == NOT_AVAILABLE_NUM:
        val = 'Caution'
    elif (row['Stress'] >= row['YS']) & (row['Stress'] < row['UTS']):
        val = 'Check'
    elif row['Stress'] >= row['UTS']:
        val = 'Fail'
    else:
        val = ''
    return val


def add_ys_uts(material, reference_file, df_stress_temperature, case: str = ''):
    df_reference = pd.read_excel(reference_file, engine='openpyxl', sheet_name=material).iloc[:, 1:]
    # print(f'mat" {material}')
    df_stress_temperature['YS'] = get_strength_by_formula(df_reference, 'YS', df_stress_temperature['Temperature'].values)
    df_stress_temperature['UTS'] = get_strength_by_formula(df_reference, 'UTS', df_stress_temperature['Temperature'].values)
    df_stress_temperature['Note'] = df_stress_temperature.apply(stress_conditions, axis=1)

    df_with_header = pd.concat([pd.DataFrame([df_stress_temperature.columns], columns=df_stress_temperature.columns), df_stress_temperature]).reset_index(drop=True)
    if case != '':
        df_case = pd.DataFrame([['Case'], [case]], columns=['Case'])
        return pd.concat([df_case, df_with_header], axis=1).reset_index(drop=True)
    else:
        return df_with_header


def get_max_value(df, time_sec):
    # 절대값 크기로 판단 (방향성 고려)
    max_nodes = [df[param].abs().idxmax() for param in df.columns]  # columns: 'Equivalent_Stress', 'Temperature' , 'Principal1' , 'Principal3'
    tuple_max = [(max_node, df.loc[max_node, param]) for param, max_node in zip(df.columns, max_nodes)]
    return [[time_sec, max_node] + df.loc[max_node, :].values.tolist() for max_node in max_nodes], tuple_max  # 0: 'Equivalent_Stress', 1: 'Temperature', 2: 'Principal1' , 3: 'Principal3'


def get_strength_by_formula(df_ref, st_type, temp_values):
    df_st = df_ref[df_ref['Properties'] == st_type]
    strength_val = []
    for temperature_val in temp_values:
        strength_formula = find_formula_by_temperature(df_st, temperature_val)
        # print(f"type: {st_type}, strength_formula: {strength_formula} temp: {temperature_val}")
        if strength_formula is not None:
            strength_val.append(eval(strength_formula.format(Temp=temperature_val)))
        else:
            strength_val.append(NOT_AVAILABLE_NUM)
    return strength_val


def find_formula_by_temperature(df, temp_val):
    for index, row in df.iterrows():
        condition_temperature = row['Temperature']
        if '~' in condition_temperature:
            temp_range = condition_temperature.split('~')
            lower = float(temp_range[0])
            upper = float(temp_range[1])
            if lower <= temp_val <= upper:
                formula = df.loc[index, 'Formula'].replace("^", "**").replace("Temp", "{Temp}")
                return formula
        else:
            if float(condition_temperature) == temp_val:
                formula = df.loc[index, 'Formula'].replace("^", "**").replace("Temp", "{Temp}")
                return formula
    return None


def find_formula_by_strain(df, strain_val):
    for index, row in df.iterrows():
        condition_strain = row['Strain']
        if '~' in condition_strain:
            strain_range = condition_strain.split('~')
            lower = float(strain_range[0])
            upper = float(strain_range[1])
            if lower <= strain_val <= upper:
                formula = df.loc[index, 'Formula'].replce("LOG", "*math.log10").replace("^", "**").replace("S", "{S}")
                return formula
        else:
            if float(condition_strain) == strain_val:
                formula = df.loc[index, 'Formula'].replce("LOG", "*math.log10").replace("^", "**").replace("S", "{S}")
                return formula
    return None


def concat_dataframes_with_gaps(df_lists, gap_rows=1):
    result_list = []
    empty_df = pd.DataFrame(np.nan, index=range(gap_rows), columns=df_lists[0].columns)
    for i, df in enumerate(df_lists):
        result_list.append(df)

        # 빈 행 추가 (마지막 그룹 제외)
        if i != len(df_lists)-1:
            result_list.append(empty_df)

    return pd.concat(result_list, axis=0).reset_index(drop=True)


def check_keyword_layout(x):
    if isinstance(x, str):
        return KEYWORD[0] in x or KEYWORD[1] in x
    return False


def combine_column_values(df):
    last_index = df.index[-1] + 1  # xlsx에서는 1부터 시작

    # 각 열의 유효한 값들을 하나의 문자열로 결합
    combined_columns = {}

    for column in df.columns:
        # 각 열의 유효한 값만 필터링하고 문자열로 변환
        valid_values = [str(val) for val in df[column] if pd.notna(val) and str(val).strip() != '']
        result = []  # 중복 제거된 값들이 들어갈 리스트
        for value in valid_values:
            if value not in result:
                if 'Compressor' not in value:  # Compressor는 풀네임에서 제외
                    result.append(value)
        if result:  # result값이 유효하다면 값들을 '_'로 묶음
            full_name = '_'.join(result).upper().replace('-', '_').replace(' ', '_')  # Ansys에서는 대문자 및 _로 표현됨
            combined_columns[(last_index, column+1)] = full_name  # xlsx에서는 1부터 시작
    return combined_columns


def split_dataframe(df, chunk_size=3):
    # 전체 행 수 확인
    total_rows = len(df)

    # 분리된 데이터프레임을 저장할 리스트
    df_list = []

    # chunk_size만큼 분리
    for i in range(0, total_rows, chunk_size):
        chunk = df.iloc[i:i + chunk_size]
        df_list.append(chunk)

    return df_list


def read_excel_with_merged_cells(ws):
    # 병합된 셀 정보 가져오기
    merged_ranges = ws.merged_cells.ranges

    # 병합된 셀의 값을 저장할 딕셔너리
    merged_values = {}

    # 병합된 셀의 값을 처리
    for merged_range in merged_ranges:
        # 병합 범위의 첫 번째 셀 값 가져오기
        first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        value = first_cell.value

        # 병합된 모든 셀에 같은 값 적용
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value

    # 데이터프레임으로 변환할 데이터 준비
    data = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            # 병합된 셀이면 저장된 값 사용, 아니면 셀 값 사용
            if (cell.row, cell.column) in merged_values:
                row_data.append(merged_values[(cell.row, cell.column)])
            else:
                row_data.append(cell.value)
        data.append(row_data)

    # 데이터프레임 생성
    df_data = pd.DataFrame(data)

    # KTP와 2D_Layout 행만 추출
    mask = df_data.iloc[:, 0].apply(check_keyword_layout)
    filtered_df = df_data[mask]
    named_dfs = split_dataframe(filtered_df, chunk_size=3)
    dict_list = [combine_column_values(named_df) for named_df in named_dfs]
    merged_dict = {}
    for d in dict_list:
        merged_dict.update(d)
    return merged_dict


def find_changed_points(data):
    diff = np.diff(data)
    sign, changed_points = None, []
    for index, diff in enumerate(diff):
        if diff != 0:
            current_sign = 'positive' if diff > 0 else 'negative'
            if sign != current_sign:
                if sign is not None:
                    changed_points.append(index)
            sign = current_sign
    return changed_points


def rainflow_point(result_path, part_name, df, changed_indices):
    # 그래프 생성
    fig, ax = plt.subplots(figsize=(10, 6))

    # 데이터프레임 열 그래프로 표시
    ax.plot(df['Time'], df['Stress'], color='black', linewidth=1)

    x_changed = [df['Time'].values[0]] + df.loc[changed_indices, 'Time'].values.tolist() + [df['Time'].values[-1]]
    y_changed = [df['Stress'].values[0]] + df.loc[changed_indices, 'Stress'].values.tolist() + [df['Stress'].values[-1]]
    # 지정된 인덱스에 X 마커 표시
    # ax.scatter(df.loc[changed_indices, 'No'], df.loc[changed_indices, 'Stress'], marker='o', color='red', s=10, linewidth=2, label='changed index')
    ax.plot(x_changed, y_changed, '--o', color='red', ms=3, linewidth=1, label='changed index')

    # 그래프 스타일 및 레이블 설정
    ax.set_title(f"Rainflow cycle of {part_name}", fontsize=14)
    ax.set_xlabel('Step', fontsize=12)
    ax.set_ylabel('Stress[MPa]', fontsize=12)
    ax.grid(True, linestyle='--', alpha=0.7)
    ax.legend(loc='upper right')
    plt.tight_layout()

    plt.savefig(f'{result_path}/{part_name}.png', format='png')  # result 폴더가 있어야 함
    plt.cla()  # clear the current axes
    plt.clf()  # clear the current figure
    plt.close()  # closes the current figure


def rainflow_counting(data):
    """
    Rainflow counting algorithm implementation

    Parameters:
    -----------
    data : array_like
        Time series data or array of reversals

    Returns:
    --------
    list
        List of tuples (range, mean, count, start_idx, end_idx)
        where range is the cycle amplitude (half the peak-to-valley distance),
        mean is the cycle mean value,
        count is always 1.0 for full cycles,
        start_idx and end_idx are the indices in the reversal array
    """
    # Create a stack to process the rainflow algorithm
    stack = deque()
    cycles = []

    # Process each reversal
    for i, point in enumerate(data):
        if len(stack) < 2:
            stack.append((i, point))
            continue

        # Process the stack when we have at least 3 points
        while len(stack) >= 2:
            # Peek the top 2 points from the stack
            idx1, point1 = stack[-2]
            idx2, point2 = stack[-1]

            # Calculate the ranges
            range_X = abs(point2 - point1)  # 구식
            range_Y = abs(point - point2)  # 최신

            # Check if we have a closed cycle
            if range_X <= range_Y:
                # Extract the cycle range and mean
                cycle_range = range_X
                cycle_mean = (point1 + point2) / 2

                # Add cycle to the result
                cycles.append((max(point1, point2), min(point1, point2), cycle_mean, cycle_range, 1.0))

                # Remove the points that formed the cycle
                stack.pop()
                stack.pop()

                # If the stack is empty, process next point
                if not stack:
                    stack.append((i, point))
                    break
            else:
                break

        # Add the current point to the stack
        stack.append((i, point))

    # Process remaining points in the stack as half-cycles
    remaining = list(stack)
    for i in range(len(remaining) - 1):
        idx1, point1 = remaining[i]
        idx2, point2 = remaining[i + 1]

        cycle_range = abs(point2 - point1)
        cycle_mean = (point1 + point2) / 2

        # Add half-cycle to the result (count as 0.5)
        cycles.append((max(point1, point2), min(point1, point2), cycle_mean, cycle_range, 1.0))

    return pd.DataFrame(cycles, columns=['Max', 'Min', 'Mean', 'Range', 'Cycle'])
