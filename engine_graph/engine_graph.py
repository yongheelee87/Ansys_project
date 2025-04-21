import pandas as pd
import numpy as np
import time
import openpyxl
import os
import matplotlib.pyplot as plt

FILE_PATH = r"./data/Mission_Test.xlsx"  # xlsx 파일 경로
COLOR_LIST = ['#808080',  # Gray
              '#3F51B5',  # Indigo
              '#03A9F4',  # Light Blue
              '#4CAF50',  # Green
              '#FFEB3B',  # Yellow
              '#FF00FF',  # Magenta
              '#F44336',  # Red
              '#00BCD4',  # Cyan
              '#FF9800'   # Orange
              '#E91E63',  # Pink
              '#2196F3',  # Blue
              '#9C27B0',  # Purple
              '#673AB7',  # Deep Purple
              '#009688',  # Teal
              '#8BC34A',  # Light Green
              '#CDDC39',  # Lime
              '#FFC107',  # Amber
              '#FF5722'  # Deep Orange
              ]

ALPHA_AREA = 0.1
ALPHA_LINE = 0.3


def isdir_and_make(dir_name: str):
    if not (os.path.isdir(dir_name)):
        os.makedirs(name=dir_name, exist_ok=True)
        print(f"Success: Create {dir_name}\n")


def remove_duplicates(lst):
    result = []
    for item in lst:
        if item not in result:
            result.append(item)
    return result


def reorganize_dataframe(df):
    # 두 번째 행을 새로운 컬럼으로 설정
    new_columns = df.iloc[1]

    # 새로운 데이터프레임 생성 (세 번째 행부터)
    new_df = df.iloc[2:].reset_index(drop=True)

    # 새로운 컬럼 이름 설정
    new_df.columns = new_columns

    return new_df


def read_excel_with_merged_cells(file_path, sheet_name: str or int = 1):
    # 워크북 로드
    wb = openpyxl.load_workbook(file_path)
    ws = wb[wb.sheetnames[sheet_name] if isinstance(sheet_name, int) else sheet_name]

    # 병합된 셀 정보 가져오기
    merged_ranges = ws.merged_cells.ranges

    # 병합된 셀의 값을 저장할 딕셔너리
    merged_values = {}

    # 병합된 셀의 값을 처리
    for merged_range in merged_ranges:
        # 병합 범위의 첫 번째 셀 값 가져오기
        first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        value = first_cell.value

        # 병합된 모든 셀에 같은 값 적용
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value

    # 데이터프레임으로 변환할 데이터 준비
    data = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            # 병합된 셀이면 저장된 값 사용, 아니면 셀 값 사용
            if (cell.row, cell.column) in merged_values:
                row_data.append(merged_values[(cell.row, cell.column)])
            else:
                row_data.append(cell.value)
        if all(x is None for x in row_data) is False:
            data.append(row_data)

    # 데이터프레임 생성
    df_data = pd.DataFrame(data).dropna(axis=1, how='all')
    part_names = [name for name in df_data.iloc[0, :].unique() if name is not None]

    df_rev = reorganize_dataframe(df_data)

    return df_rev, part_names


def create_stress_temp_plot(x_data, df, dict_changes, part, time_graph):
    # Stress와 Temp가 포함된 열 찾기
    stress_columns = [col for col in df.columns if 'Stress' in col]
    temp_columns = [col for col in df.columns if 'Temp' in col]

    # 그래프 생성
    fig, ax1 = plt.subplots(figsize=(12, 8))

    # 왼쪽 y축 (Stress)
    for col in stress_columns:
        ax1.plot(x_data, df[col], '-', label=col)
        ax1.set_ylabel(col)
    ax1.set_xlabel('Time[s]')
    ax1.set_xlim(0, x_data.max())  # x축 범위를 데이터의 최소값과 최대값으로 설정 x_data.min(), x_data.max()
    ax1.tick_params(axis='y', labelcolor='tab:blue')

    # 오른쪽 y축 (Temperature)
    ax2 = ax1.twinx()
    for col in temp_columns:
        ax2.plot(x_data, df[col], '-', color='tab:red', label=col)
        ax2.set_ylabel(col)
    ax2.tick_params(axis='y', labelcolor='tab:red')

    # 범례 설정
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')

    # 영역 구분
    locations = []
    ratings = list(dict_changes.values())
    color_rate = {rate: color for rate, color in zip(remove_duplicates(ratings), COLOR_LIST)}
    for indices, rating in dict_changes.items():
        ax1.axvspan(x_data[indices[0]], x_data[indices[1]], alpha=ALPHA_AREA, color=color_rate[rating])
        if indices[0] != 0:
            ax1.axvline(x=x_data[indices[0]], color='black', linestyle='--', linewidth=1, alpha=ALPHA_LINE)
        locations.append(np.mean([x_data[indices[0]], x_data[indices[1]]]))

    ax_twin = ax1.twiny()
    ax_twin.set_xlim(ax1.get_xlim())
    ax_twin.set_xticks(np.array(locations))
    ax_twin.tick_params(axis='x', length=0)  # x축의 tick length를 0으로 설정
    ax_twin.set_xticklabels(ratings)

    plt.title(f'[{name}] Stress and Temperature', fontsize=18, pad=15)
    plt.savefig(f'./result/engine_graph_{time_graph}/{part}.png', format='png')  # result 폴더가 있어야 함
    plt.cla()  # clear the current axes
    plt.clf()  # clear the current figure
    plt.close()  # closes the current figure


def find_string_changes(df, column_name):
    """
    데이터프레임의 문자열 컬럼에서 값이 변하는 지점을 찾아
    변화가 발생한 문자열과 해당 인덱스를 반환합니다.

    Parameters:
    df (pandas.DataFrame): 분석할 데이터프레임
    column_name (str): 분석할 문자열 컬럼명

    Returns:
    tuple: (변화가 발생한 인덱스 리스트, 변화가 발생한 문자열 리스트)
    """
    # 이전 값과 다른 값을 가진 행 찾기
    value_changes = df[column_name] != df[column_name].shift()

    # 첫 번째 행은 항상 포함 (초기값)
    value_changes.iloc[0] = True

    # 변화가 발생한 지점의 인덱스와 문자열 추출
    change_indices = df.index[value_changes].tolist()
    change_indices.append(df.index[-1])
    change_values = df.loc[value_changes, column_name].tolist()

    return {(change_indices[i], change_indices[i+1]): change_values[i] for i in range(len(change_indices) - 1)}


if __name__ == "__main__":
    df_excel, parts = read_excel_with_merged_cells(file_path=FILE_PATH, sheet_name='Sheet2')
    temp_indices = [i for i, col in enumerate(df_excel.columns) if 'Temp' in col]
    exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
    isdir_and_make(f'./result/engine_graph_{exe_time}')  # result 결과 시간 폴더 생성
    start_index = 0
    for index, name in zip(temp_indices, parts):
        # 첫 번째 열을 x축 데이터로 설정
        time_data = df_excel.iloc[:, 0]
        create_stress_temp_plot(x_data=time_data,
                                df=df_excel.iloc[:, start_index:index + 1],
                                dict_changes=find_string_changes(df_excel, df_excel.columns[1]),
                                part=name,
                                time_graph=exe_time)
        start_index = index+1
