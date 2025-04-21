import numpy as np
from ansys.dpf import core as dpf
from ansys.dpf.core import operators as ops
import os
import openpyxl
import time
import yaml
import pandas as pd
from functions import (COLUMNS_TOTAL, get_max_dataframe, extract_all_stress_temperature,
                       get_file_name_modification_date, isdir_and_make, add_ys_uts,
                       concat_dataframes_with_gaps, read_excel_with_merged_cells, rainflow_cycle,
                       rainflow_point, extract_yellow_cells_info, reorder_data)


ROOT_PATH = os.getcwd()  # 현재 위치 가져오기
RESULT_PATH = f"{ROOT_PATH}/result"
DATA_PATH = f"{ROOT_PATH}/data"
REFERENCE_EXCEL = f"{DATA_PATH}/Calculation_Reference.xlsx"
RESULT_EXCEL = f"{DATA_PATH}/Result_Format.xlsx"
ABBREVIATION = f"{DATA_PATH}/abbreviation.yaml"
NAME_SELECTION_PREFIX = "S_"
COLUMNS = ["Time", "NodeID", "Stress", "Temperature"]
RESULT_COLUMNS = ["Part", "Time", "Max_Temp", "NodeID_T", "Stress", "NodeID_S", "YS", "UTS", "Temperature", "Material"]
CASES = ['Max EQ of Each Step', 'Max Temp of Each Step', 'Max EQ', 'Max Temp', 'Max Principal', 'Min Principal', 'Signed Max EQ']
RAINFLOW_ONOFF = True


class DPF_Stress_Temperature:
    def __init__(self, save_data, rst_path):
        isdir_and_make(RESULT_PATH)  # result 폴더 생성
        self.save_data = save_data
        self.abbre_mat = None
        self.model = None
        self.stress_op = None
        self.temperature_op = None
        self.mesh = None
        self.metadata = None
        self.bodies = []
        self.result_part = {}
        self.result_time = []
        self.rst_path = rst_path
        self.sheet_name = []
        self.time_sec = []
        self.step_ids = []
        self.exe_time = ''
        # self.coordinates = {}

    def all_data_to_parquet(self):
        print("************************************************************")
        print("*** #0.Mode DPF Stress Temperature Data to Parquet with All nodes Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')

        parquet_path = get_file_name_modification_date(self.rst_path)
        if not os.path.isdir(f'{DATA_PATH}/{parquet_path}'):
            self._load_abbreviation_material()  # 약자 읽기
            self._load_rst()  # model, rst 읽기

            self._load_part()  # S_ part 읽기
            self.set_operation()  # operation 준비

            self._extract_all_nodes_data(data_path=f'{DATA_PATH}/{parquet_path}')
        else:
            print(f"*** [STOP!!] {parquet_path} already exists\n")

    def analyze_base_model(self):
        print("************************************************************")
        print("*** #1.Mode DPF Stress Temperature Analysis Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')
        if self.save_data is True:
            isdir_and_make(f'{RESULT_PATH}/{self.exe_time}')  # result 결과 시간 폴더 생성
        isdir_and_make(f'{RESULT_PATH}/Mode 1')  # result 결과 시간 폴더 생성

        parquet_path = get_file_name_modification_date(self.rst_path)
        if os.path.isdir(f'{DATA_PATH}/{parquet_path}'):
            self._load_result_format()
            self.write_result_xlsx(self.wrap_result(data_path=f'{DATA_PATH}/{parquet_path}'))
        else:
            print(f"*** [STOP!!] {parquet_path} does NOT exist\n*** Please Check if the most recently modified result is correct\n")
        self.stop('#1.Mode')
        
    def check_temperature_range(self):
        print("************************************************************")
        print("*** #2.Mode Find Temperature Range Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')
        isdir_and_make(f'{RESULT_PATH}/Mode 2')  # result 폴더 생성

        parquet_path = get_file_name_modification_date(self.rst_path)
        if os.path.isdir(f'{DATA_PATH}/{parquet_path}'):
            self.wrap_temperature_result(data_path=f'{DATA_PATH}/{parquet_path}')
        else:
            print(f"*** [STOP!!] {parquet_path} does NOT exist\n*** Please Check if the most recently modified result is correct\n")
        self.stop('#2.Mode')

    def analyze_base_model_all_nodes(self):
        print("************************************************************")
        print("*** #3.Mode DPF Stress Temperature Analysis with All nodes Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')

        isdir_and_make(f'{RESULT_PATH}/Mode 3')  # result 폴더 생성

        parquet_path = get_file_name_modification_date(self.rst_path)
        if os.path.isdir(f'{DATA_PATH}/{parquet_path}'):
            self.wrap_all_nodes_result(data_path=f'{DATA_PATH}/{parquet_path}')
        else:
            print(f"*** [STOP!!] {parquet_path} does NOT exist\n*** Please Check if the most recently modified result is correct\n")

        self.stop('#3.Mode')

    def analyze_base_model_specific_nodes(self, nodes):
        print("************************************************************")
        print("*** #4.Mode DPF Stress Temperature Analysis with Specific nodes Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')

        isdir_and_make(f'{RESULT_PATH}/Mode 4')  # result 폴더 생성

        parquet_path = get_file_name_modification_date(self.rst_path)
        if os.path.isdir(f'{DATA_PATH}/{parquet_path}'):
            self.wrap_specific_nodes_result(node_ids=nodes, data_path=f'{DATA_PATH}/{parquet_path}')

        self.stop('#4.Mode')

    def analyze_static(self):
        print("************************************************************")
        print("*** #5.Mode DPF Stress Temperature Static Analysis Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')

        isdir_and_make(f'{RESULT_PATH}/Mode 5')  # result 폴더 생성

        parquet_path = get_file_name_modification_date(self.rst_path)
        if os.path.isdir(f'{DATA_PATH}/{parquet_path}'):
            self.wrap_layout_analysis(method='Static', data_path=f'{DATA_PATH}/{parquet_path}')
        self.stop('#5.Mode')

    def analyze_transient(self):
        print("************************************************************")
        print("*** #6.Mode DPF Stress Temperature Transient Analysis Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')

        isdir_and_make(f'{RESULT_PATH}/Mode 6')  # result 폴더 생성

        parquet_path = get_file_name_modification_date(self.rst_path)
        if os.path.isdir(f'{DATA_PATH}/{parquet_path}'):
            self.wrap_layout_analysis(method='Transient', data_path=f'{DATA_PATH}/{parquet_path}')
        self.stop('#6.Mode')

    def stop(self, mode: str):
        print(f'Ending at: {time.strftime("%a, %d-%b-%Y %I:%M:%S", time.localtime(time.time()))}')
        print("************************************************************")
        print(f"*** {mode} DPF Stress Temperature Analysis completed")
        print("************************************************************\n")

    def _load_part(self):
        # Body 가져오기
        self.bodies = [body_name for body_name in self.metadata.available_named_selections if NAME_SELECTION_PREFIX in body_name[:2]]
        print(f"[INFO] Selected Names: {self.bodies}\n")

        # self.coordinates = {node.id: node.coordinates for node in self.mesh.nodes}

    def set_operation(self):
        # time_frequencies 객체 가져오기
        time_freq = self.metadata.time_freq_support.time_frequencies
        self.time_sec = time_freq.data
        self.step_ids = time_freq.scoping.ids

        self.stress_op = self.model.results.stress()
        self.stress_op.inputs.requested_location.connect('Nodal')

        self.temperature_op = self.model.results.structural_temperature()
        self.temperature_op.inputs.requested_location.connect('Nodal')

    def wrap_result(self, data_path):
        # 결과 경로 변수화
        result_dir = f'{RESULT_PATH}/Mode 1/{self.exe_time}'
        isdir_and_make(result_dir)

        # 파트 목록 추출
        self.bodies = [f.replace('.parquet', '') for f in os.listdir(data_path) if '.parquet' in f]

        # 메인 결과 파일 경로
        main_result_file = f'{result_dir}/result_stress_temp.xlsx'

        # 레인플로우 분석 관련 파일 및 Writer 객체 준비
        writers = {'main': pd.ExcelWriter(main_result_file, engine='openpyxl')}

        if RAINFLOW_ONOFF:
            # 레인플로우 관련 파일 설정
            rainflow_files = {
                'rainflow': 'result_rainflow_point.xlsx',
                'signed_rainflow': 'result_signed_rainflow_point.xlsx',
                'cycle': 'result_cycle_rainflow_point.xlsx',
                'cycle_signed': 'result_cycle_signed_rainflow_point.xlsx'
            }

            # 각 파일에 대한 Writer 객체 생성
            for key, filename in rainflow_files.items():
                writers[key] = pd.ExcelWriter(f'{result_dir}/{filename}', engine='openpyxl')

        # 요약 데이터를 위한 리스트 (최종 크기 미리 할당)
        estimated_rows = len(self.bodies) * len(self.result_time)
        summary_data = [None] * estimated_rows
        row_idx = 0

        # 각 파트별 처리
        for part in self.bodies:
            # 현재 파트 데이터 필터링
            df_part = pd.read_parquet(f'{data_path}/{part}.parquet')
            material = df_part['Material'].iloc[0]

            # 스트레스 및 온도 데이터 추출 및 YS/UTS 추가
            stress_temp_data = extract_all_stress_temperature(df=df_part)
            lst_df_st_temp_rev = []

            # 각 케이스별 YS/UTS 추가 (리스트 컴프리헨션 대신 루프 사용)
            for idx, case in enumerate(CASES):
                lst_df_st_temp_rev.append(add_ys_uts(material, REFERENCE_EXCEL, stress_temp_data[idx], case))

            # 레인플로우 분석 실행 (해당하는 경우)
            if RAINFLOW_ONOFF:
                self._run_rainflow_point(df=lst_df_st_temp_rev[2], part=part,
                                         cylcle=writers['cycle'],
                                         rainflow=writers['rainflow'],
                                         signed=False)
                self._run_rainflow_point(df=lst_df_st_temp_rev[-1], part=part,
                                         cylcle=writers['cycle_signed'],
                                         rainflow=writers['signed_rainflow'],
                                         signed=True)

            # 각 시간별 최대 스트레스 및 온도 데이터 처리
            for t in self.result_time:
                # 해당 시간의 최대 스트레스 및 온도 데이터 가져오기
                stress_time_data = lst_df_st_temp_rev[0][lst_df_st_temp_rev[0]['Time'] == t]
                temp_time_data = lst_df_st_temp_rev[1][lst_df_st_temp_rev[1]['Time'] == t]

                # 데이터가 존재하는지 확인
                if len(stress_time_data) > 0 and len(temp_time_data) > 0:
                    max_stress_by_time = stress_time_data.iloc[0]
                    max_temperature_by_time = temp_time_data.iloc[0]

                    # 요약 데이터 추가
                    summary_data[row_idx] = [
                        part, t,
                        max_temperature_by_time[4], max_temperature_by_time[2],  # 온도, 노드 ID
                        max_stress_by_time[3], max_stress_by_time[2],  # 스트레스, 노드 ID
                        max_stress_by_time[5], max_stress_by_time[6],  # YS, UTS
                        max_stress_by_time[4], material  # 온도, 재료
                    ]
                    row_idx += 1

            # 데이터프레임 병합 및 엑셀 시트 작성
            df_st_temp = concat_dataframes_with_gaps(lst_df_st_temp_rev)
            df_st_temp.to_excel(writers['main'], sheet_name=part, index=False, header=False)

        # 사용하지 않은 공간 제거
        summary_data = [row for row in summary_data if row is not None]

        # 모든 Writer 객체 닫기
        for writer in writers.values():
            writer.close()

        # 요약 데이터 반환
        return pd.DataFrame(summary_data, columns=RESULT_COLUMNS)

    def write_result_xlsx(self, df):
        # 파일 경로 변수화
        result_file = f'{RESULT_PATH}/Mode 1/{self.exe_time}/result_summary.xlsx'

        # 필요한 데이터만 미리 준비 (시간별로 그룹화)
        time_df_dict = {}

        # 워크북 로드
        wb = openpyxl.load_workbook(RESULT_EXCEL)

        # 모든 시트에 대해 처리
        for sh, parts in self.result_part.items():
            ws = wb[sh]

            # 시간 정보 추출 (한 번만 처리)
            time_str = sh.split('_')[0].replace('s', '')
            time_val = int(time_str)

            # 해당 시간에 대한 데이터프레임 캐싱 (중복 필터링 방지)
            if time_val not in time_df_dict:
                time_df_dict[time_val] = df[df['Time'] == time_val]
            df_time = time_df_dict[time_val]

            # Temperature 시트인지 확인
            is_temp_sheet = 'Temp' in sh.split('_')[-1]

            # 파트별 처리
            for pos, part in parts.items():
                # 첫번째 열은 KTP 포함 제외
                if pos[1] == 1:
                    continue

                # 해당 파트 데이터 가져오기 (인덱싱 미리 계산)
                part_rows = df_time['Part'] == part
                if not part_rows.any():
                    continue  # 해당 파트가 없으면 건너뛰기

                part_data = df_time[part_rows].iloc[0]

                # 행 오프셋 계산 (반복 더하기 연산 제거)
                base_row = pos[0]

                if is_temp_sheet:
                    # Temperature 데이터 입력
                    ws.cell(row=base_row + 1, column=pos[1]).value = part_data[2]  # Temperature
                    ws.cell(row=base_row + 2, column=pos[1]).value = part_data[3]  # Node ID
                else:
                    # Stress 데이터 입력
                    ws.cell(row=base_row + 1, column=pos[1]).value = part_data[9]  # 소재
                    ws.cell(row=base_row + 2, column=pos[1]).value = part_data[6]  # YS
                    ws.cell(row=base_row + 3, column=pos[1]).value = part_data[7]  # UTS
                    ws.cell(row=base_row + 4, column=pos[1]).value = part_data[4]  # Stress
                    ws.cell(row=base_row + 5, column=pos[1]).value = part_data[8]  # Temperature
                    ws.cell(row=base_row + 6, column=pos[1]).value = part_data[5]  # Node ID

        # 파일 저장
        wb.save(result_file)
        wb.close()

    def wrap_temperature_result(self, data_path):
        # 파트 목록 저장
        self.bodies = [f.replace('.parquet', '') for f in os.listdir(data_path) if '.parquet' in f]

        # 결과 저장용 딕셔너리 직접 생성 (불필요한 리스트 및 중간 변환 제거)
        result_data = {
            '파트명': [],
            '최소온도': [],
            '최대온도': []
        }

        # 각 파트별 처리
        for part in self.bodies:
            df_part = pd.read_parquet(f'{data_path}/{part}.parquet')

            # 결과 딕셔너리에 직접 추가
            result_data['파트명'].append(part)  # df_part 대신 part 이름 저장
            result_data['최소온도'].append(df_part["Temperature"].min())
            result_data['최대온도'].append(df_part["Temperature"].max())

        # 파일 경로
        output_file = f'{RESULT_PATH}/Mode 2/{self.exe_time}_result_temperature.xlsx'

        # with 문을 사용하여 파일 자동 닫기
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            pd.DataFrame(result_data).to_excel(writer, sheet_name='온도', index=False)

    def wrap_all_nodes_result(self, data_path):
        result_dir = f'{RESULT_PATH}/Mode 3/{self.exe_time}'
        isdir_and_make(result_dir)  # result 폴더 생성

        # 파트 목록 한 번만 계산
        self.bodies = [f.replace('.parquet', '') for f in os.listdir(data_path) if '.parquet' in f]

        # 1. 요약 파일 생성
        summary_file = f'{result_dir}_result_stress_temp_all_nodes_summary.xlsx'

        with pd.ExcelWriter(summary_file, engine='openpyxl') as writer_sum:
            # 2. 각 파트별 처리
            for part in self.bodies:
                # 현재 파트의 데이터 필터링 (한 번만 수행)
                df_part = pd.read_parquet(f'{data_path}/{part}.parquet')
                material = df_part['Material'].iloc[0]

                # 열 이름 변경 (필요한 경우)
                if 'Equivalent_Stress' in df_part.columns:
                    df_part = df_part.rename(columns={'Equivalent_Stress': 'Stress'})

                #  YS UTS check 추가
                df_st_temp = add_ys_uts(material, REFERENCE_EXCEL, df_part, case='')

                # 파트별 엑셀 파일 생성
                part_file = f'{result_dir}/{part}.xlsx'

                # 요약 데이터 준비
                summary_data = {'Time': [], 'Check_Nodes': [], 'Fail_Nodes': []}

                with pd.ExcelWriter(part_file, engine='openpyxl') as writer:
                    # 시간별 데이터 처리
                    for t, df_time in df_st_temp.groupby('Time'):
                        # 체크 및 실패 노드 필터링 (불필요한 리스트 변환 없이)
                        check_mask = df_time['Note'] == 'Check'
                        fail_mask = df_time['Note'] == 'Fail'

                        # 노드 ID 추출 및 문자열 포맷팅
                        check_nodes = ", ".join(map(str, df_time.loc[check_mask, 'NodeID'].values)) if check_mask.any() else ''
                        fail_nodes = ", ".join(map(str, df_time.loc[fail_mask, 'NodeID'].values)) if fail_mask.any() else ''

                        # 요약 데이터 업데이트
                        summary_data['Time'].append(t)
                        summary_data['Check_Nodes'].append(check_nodes)
                        summary_data['Fail_Nodes'].append(fail_nodes)

                        # 시간별 시트 작성
                        df_time.to_excel(writer, sheet_name=f'{int(t)}s', index=False)

                # 요약 시트 작성
                pd.DataFrame(summary_data).to_excel(writer_sum, sheet_name=part, index=False)

    def wrap_specific_nodes_result(self, node_ids: str, data_path: str):
        # 결과 파일 경로 설정
        result_file = f'{RESULT_PATH}/Mode 4/{self.exe_time}_result_stress_temp_specific_nodes.xlsx'

        # 노드 ID 리스트 변환 (한 번에 처리)
        lst_node = [int(node.strip()) for node in node_ids.split(',')]

        # 각 노드별 데이터프레임 및 처리 결과 저장
        node_dataframes = {}

        lst_df = [pd.read_parquet(f'{data_path}/{part}') for part in [f for f in os.listdir(data_path) if '.parquet' in f]]
        df_data = pd.concat(lst_df, ignore_index=True)

        # 각 노드별 처리
        for node in lst_node:
            # 노드별 데이터 필터링
            df_node = df_data[df_data['NodeID'] == node]

            # 노드 데이터가 있는지 확인
            if df_node.empty:
                continue

            material = df_node['Material'].iloc[0]

            # 열 이름 변경 (필요한 경우)
            if 'Equivalent_Stress' in df_node.columns:
                df_node = df_node.rename(columns={'Equivalent_Stress': 'Stress'})

            # YS UTS check 추가
            df_st_temp = add_ys_uts(material, REFERENCE_EXCEL, df_node, case='')
            node_dataframes[node] = df_st_temp

        # 모든 노드 데이터가 있는지 확인
        if not node_dataframes:
            print(f"경고: 지정된 노드 ID {node_ids}에 해당하는 데이터가 없습니다.")
            return

        # Excel 파일 작성
        with pd.ExcelWriter(result_file, engine='openpyxl') as writer:
            # 전체 데이터를 합쳐서 'data' 시트에 작성
            df_node_st_temp = concat_dataframes_with_gaps(list(node_dataframes.values()))
            df_node_st_temp.to_excel(writer, sheet_name='data', index=False, header=True)

            # 각 노드별 데이터를 개별 시트에 작성
            for node, df in node_dataframes.items():
                df.to_excel(writer, sheet_name=f'node {node}', index=False, header=True)

    def _load_abbreviation_material(self):
        with open(ABBREVIATION, encoding='utf-8-sig') as f:
            setting = yaml.load(f, Loader=yaml.SafeLoader)
        self.abbre_mat = setting['Material']

    def _load_rst(self):
        # rst 파일 찾기
        rst_file = [file for file in os.listdir(self.rst_path) if '.rst' in file]
        # rst 파일 직접 로드
        self.model = dpf.Model(os.path.join(self.rst_path, rst_file[0]))
        self.metadata = self.model.metadata
        self.mesh = self.metadata.meshed_region
        print(self.model)

    def _load_result_format(self):
        # 워크북 로드
        wb = openpyxl.load_workbook(RESULT_EXCEL)
        parts, list_time_in = {}, []
        for sheet_name in wb.sheetnames:
            list_time_in.append(int(sheet_name.split('_')[0].replace('s', '')))
            parts[sheet_name] = read_excel_with_merged_cells(wb[sheet_name])
        self.result_time = list(set(list_time_in))
        self.result_part = parts

    def _run_rainflow_point(self, df, part, cylcle, rainflow, signed: bool = False):
        df_cycle, df_max, changed_indices = rainflow_cycle(df)
        # 엑셀 출력 작업
        df_cycle.to_excel(cylcle, sheet_name=part, index=False)

        df_max.loc[changed_indices].to_excel(rainflow, sheet_name=part, index=False)

        # 경로 문자열 포맷팅 개선
        result_dir = f'{RESULT_PATH}/Mode 1/{self.exe_time}'

        name = f'{part}_Signed' if signed is True else part
        rainflow_point(result_path=result_dir, part_name=name, df=df_max, changed_indices=changed_indices)

    def _extract_all_nodes_data(self, data_path):
        isdir_and_make(data_path)

        # 파트별 처리
        for body in self.bodies:
            # 파트 데이터프레임을 위한 리스트
            part_data = []
            
            # 파트 정보 추출
            part_name = '_'.join(body.split('_')[2:])
            material = self.abbre_mat[body.split('_')[1]]

            # 메시 스코핑 설정 (한 번만)
            mesh_scoping = self.metadata.named_selection(body)
            self.stress_op.inputs.mesh_scoping.connect(mesh_scoping)
            self.temperature_op.inputs.mesh_scoping.connect(mesh_scoping)

            time_index = 0

            # 각 스텝 처리
            for step in self.step_ids:
                # 시간 스코핑 설정
                time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)

                # 응력 및 온도 연산을 위한 스코핑 설정 (한 번만)
                self.stress_op.inputs.time_scoping(time_steps_scoping)
                self.temperature_op.inputs.time_scoping(time_steps_scoping)

                # 응력 연산
                eqv_op = ops.invariant.von_mises_eqv_fc(self.stress_op)
                eqv_field = eqv_op.outputs.fields_container()

                # 주응력 연산
                principal_op = ops.invariant.principal_invariants_fc(self.stress_op)
                principal_stress_1 = principal_op.outputs.fields_eig_1()
                principal_stress_3 = principal_op.outputs.fields_eig_3()

                # 온도 필드 가져오기
                temperature_field = self.temperature_op.outputs.fields_container()

                # 필드 데이터 처리
                for eqv_stress, temperature, p_stress_1, p_stress_3 in zip(eqv_field, temperature_field,
                                                                           principal_stress_1, principal_stress_3):
                    # 정수 시간에 대해서만 처리
                    if float(self.time_sec[time_index]).is_integer():
                        # 공통 인덱스 사용하여 데이터프레임 구성
                        node_ids = eqv_stress.scoping.ids

                        reorder_temperature = reorder_data(node_ids, temperature.scoping.ids, temperature.data)

                        # 데이터 딕셔너리 구성 (더 효율적인 방법)
                        data_dict = {
                            'Equivalent_Stress': eqv_stress.data,
                            'Temperature': reorder_temperature,
                            'Principal_1': p_stress_1.data,
                            'Principal_3': p_stress_3.data,
                            'Part': part_name,
                            'Time': round(self.time_sec[time_index], 1),
                            'NodeID': node_ids,
                            'Material': material
                        }

                        # 데이터프레임 생성 및 필요한 열만 선택
                        df_st_temp = pd.DataFrame(data_dict)
                        df_st_temp = df_st_temp[COLUMNS_TOTAL]

                        # 파트 결과 리스트에 추가
                        part_data.append(df_st_temp)

                    time_index += 1
            pd.concat(part_data, ignore_index=True).to_parquet(f'{data_path}/{part_name}.parquet')
            # pd.concat(part_data, ignore_index=True).to_csv(f'{data_path}/{part_name}.csv')

    def wrap_layout_analysis(self, method, data_path):
        if method == 'Static': # 결과 경로 변수화
            result_format = f"{DATA_PATH}/2D_Layout_Result_Static.xlsx"
            result_file = f'{RESULT_PATH}/Mode 5/{self.exe_time}_2d_layout_result_static.xlsx'
        else:
            result_format = f"{DATA_PATH}/2D_Layout_Result_Transient.xlsx"
            result_file = f'{RESULT_PATH}/Mode 6/{self.exe_time}_2d_layout_result_transient.xlsx'

        # 파트 목록 추출
        self.bodies = [f.replace('.parquet', '') for f in os.listdir(data_path) if '.parquet' in f]

        # 노란색 셀 정보 추출 (한 번만 수행)
        yellow_cells_info = extract_yellow_cells_info(result_format)

        # 템플릿 엑셀 파일 로드
        wb = openpyxl.load_workbook(result_format)

        # 소스 시트 참조 저장 (반복 참조 방지)
        source_sheet = wb.worksheets[0]

        # 각 파트별 처리
        for part in self.bodies:
            # 현재 파트 데이터 필터링
            df_part = pd.read_parquet(f'{data_path}/{part}.parquet')

            # 빈 데이터프레임 확인
            if df_part.empty:
                continue

            material = df_part['Material'].iloc[0]

            # 최대 응력과 온도 데이터 계산
            df_max_stress = get_max_dataframe(df_part, "Equivalent_Stress")
            df_max_temperature = df_part.loc[[df_part["Temperature"].idxmax()]]
            max_node = df_max_stress['NodeID'].iloc[0]

            # 열 이름 변경 (단일 연산으로)
            rename_cols = {'Equivalent_Stress': 'Stress'}
            df_max_stress = df_max_stress.rename(columns=rename_cols)
            df_max_temperature = df_max_temperature.rename(columns=rename_cols)

            # YS UTS check 추가
            df_max_stress_rev = add_ys_uts(material, REFERENCE_EXCEL, df_max_stress, case='')
            df_max_temperature_rev = add_ys_uts(material, REFERENCE_EXCEL, df_max_temperature, case='')

            # 최대 응력 및 안전율 계산
            max_stress = df_max_stress_rev.loc[df_max_stress_rev['Stress'].idxmax()]
            ms = (max_stress['YS'] / max_stress['Stress']) - 1
            # 레인플로우 분석 실행
            df_cycle, df_max, changed_indices = rainflow_cycle(df_max_stress_rev)

            # 시트 생성 또는 재사용
            if part in wb.sheetnames:
                ws = wb[part]
            else:
                ws = wb.copy_worksheet(source_sheet)
                ws.title = part

            # 파트 이름 설정
            ws.cell(row=1, column=1).value = part

            # 각 행의 정보 처리
            for sheet_name, rows_info in yellow_cells_info.items():
                for row in rows_info:
                    key = row['key']
                    yellow_cells = row['yellow']

                    # 필수 값 검증
                    if not yellow_cells:
                        continue

                    # 시간값 (숫자) 처리
                    if isinstance(key, (int, float)):
                        # 해당 시간의 데이터 필터링
                        time_data = df_max_stress_rev[df_max_stress_rev['Time'] == key]
                        if time_data.empty:
                            continue

                        # 스트레스 값 채우기 (첫 번째 노란색 셀)
                        ws.cell(row=yellow_cells[0][0], column=yellow_cells[0][1]).value = time_data['Stress'].iloc[0]  # 스트레스

                        # 온도, MOS 값 채우기 (두 번째 노란색 셀)
                        ws.cell(row=yellow_cells[1][0], column=yellow_cells[1][1]).value = time_data['Temperature'].iloc[0]  # 온도값
                        ws.cell(row=yellow_cells[2][0], column=yellow_cells[2][1]).value = (time_data['YS'].iloc[0]/time_data['Stress'].iloc[0])-1  # Margin of Safety

                    # 문자열 키워드 처리
                    elif isinstance(key, str):
                        # 노란색 셀 위치 추출
                        y_row, y_col = yellow_cells[0][0], yellow_cells[0][1]

                        # 키워드별 처리
                        if 'Location' in key:
                            ws.cell(row=y_row, column=y_col).value = max_node
                        elif 'Material' in key:
                            ws.cell(row=y_row, column=y_col).value = material
                        elif 'YS [Mpa] at MAX EQ' in key:
                            ws.cell(row=y_row, column=y_col).value = max_stress['YS']
                            ws.cell(row=y_row, column=y_col + 1).value = f"@{round(max_stress['Temperature'], 1)}C"
                        elif 'UTS [Mpa] at MAX EQ' in key:
                            ws.cell(row=y_row, column=y_col).value = max_stress['UTS']
                            ws.cell(row=y_row, column=y_col + 1).value = f"@{round(max_stress['Temperature'], 1)}C"
                        elif 'YS [Mpa] at MAX Temp' in key:
                            ws.cell(row=y_row, column=y_col).value = df_max_temperature_rev['YS'].iloc[0]
                            ws.cell(row=y_row,
                                    column=y_col + 1).value = f"@{round(df_max_temperature_rev['Temperature'].iloc[0], 1)}C"
                        elif 'UTS [Mpa] at MAX Temp' in key:
                            ws.cell(row=y_row, column=y_col).value = df_max_temperature_rev['UTS'].iloc[0]
                            ws.cell(row=y_row,
                                    column=y_col + 1).value = f"@{round(df_max_temperature_rev['Temperature'].iloc[0], 1)}C"
                        elif 'Mechanical' in key:
                            ws.cell(row=y_row, column=y_col).value = max_stress['Stress']
                        elif 'MS' in key:
                            ws.cell(row=y_row, column=y_col).value = ms

        # 첫번째 템플릿 시트 제거 (마지막에 한 번만)
        wb.remove(wb[wb.sheetnames[0]])

        # 결과 저장
        wb.save(result_file)
