import os
import openpyxl
import time
import numpy as np
import pandas as pd
from scipy import signal
import matplotlib.pyplot as plt
from collections import deque


NOT_AVAILABLE_NUM = -999
KEYWORD = ['프로', '2D Layout']
COLUMNS_BASIC = ["Time", "NodeID", "Equivalent_Stress", "Temperature"]
COLUMNS_ORDER = ["Equivalent_Stress", "Temperature", "Principal_1", "Principal_3"]
COLUMNS_TOTAL = ["Part", "Time", "NodeID", "Equivalent_Stress", "Principal_1", "Principal_3", "Temperature", "Material"]
CELL_COLOR = "FFFFFFCC"


def isdir_and_make(dir_name: str):
    if not (os.path.isdir(dir_name)):
        os.makedirs(name=dir_name, exist_ok=True)
        print(f"Success: Create {dir_name}\n")


def get_file_name_modification_date(file_path):
    # 파일의 마지막 수정 시간 얻기 (초 단위 timestamp)
    mod_time_timestamp = os.path.getmtime(file_path)
    # timestamp를 str객체로 변환
    mod_time_str = time.strftime('%Y%m%d_%H%M%S', time.localtime(mod_time_timestamp))
    file_source = '_'.join(file_path.split('\\')[-2:])
    # 원하는 형식으로 출력
    return f'{file_source}_{mod_time_str}'


def stress_conditions(row):
    if row['YS'] == NOT_AVAILABLE_NUM and row['UTS'] == NOT_AVAILABLE_NUM:
        val = 'N/A'
    elif row['YS'] == NOT_AVAILABLE_NUM or row['UTS'] == NOT_AVAILABLE_NUM:
        val = 'Caution'
    elif (row['Stress'] >= row['YS']) & (row['Stress'] < row['UTS']):
        val = 'Check'
    elif row['Stress'] >= row['UTS']:
        val = 'Fail'
    else:
        val = ''
    return val


def add_ys_uts(material, reference_file, df_stress_temperature, case: str = ''):
    df_reference = pd.read_excel(reference_file, engine='openpyxl', sheet_name=material).iloc[:, 1:]
    # print(f'mat" {material}')
    df_stress_temperature['YS'] = get_strength_by_formula(df_reference, 'YS', df_stress_temperature['Temperature'].values)
    df_stress_temperature['UTS'] = get_strength_by_formula(df_reference, 'UTS', df_stress_temperature['Temperature'].values)
    df_stress_temperature['Note'] = df_stress_temperature.apply(stress_conditions, axis=1)

    if case != '':
        df_with_header = pd.concat([pd.DataFrame([df_stress_temperature.columns], columns=df_stress_temperature.columns), df_stress_temperature]).reset_index(drop=True)
        return pd.concat([pd.DataFrame([['Case'], [case]], columns=['Case']), df_with_header], axis=1).reset_index(drop=True)
    else:
        return df_stress_temperature


def get_max_value(df, time_sec):
    # 절대값 크기로 판단 (방향성 고려)
    max_nodes = [df[param].abs().idxmax() for param in df.columns]  # columns: 'Equivalent_Stress', 'Temperature' , 'Principal1' , 'Principal3'
    tuple_max = [(max_node, df.loc[max_node, param]) for param, max_node in zip(df.columns, max_nodes)]
    return [[time_sec, max_node] + df.loc[max_node, :].values.tolist() for max_node in max_nodes], tuple_max  # 0: 'Equivalent_Stress', 1: 'Temperature', 2: 'Principal1' , 3: 'Principal3'


def get_strength_by_formula(df_ref, st_type, temp_values):
    df_st = df_ref[df_ref['Properties'] == st_type]
    strength_val = []
    for temperature_val in temp_values:
        strength_formula = find_formula_by_temperature(df_st, temperature_val)
        # print(f"type: {st_type}, strength_formula: {strength_formula} temp: {temperature_val}")
        if strength_formula is not None:
            strength_val.append(eval(strength_formula.format(Temp=temperature_val)))
        else:
            strength_val.append(NOT_AVAILABLE_NUM)
    return strength_val


def find_formula_by_temperature(df, temp_val):
    for index, row in df.iterrows():
        condition_temperature = row['Temperature']
        if '~' in condition_temperature:
            temp_range = condition_temperature.split('~')
            lower = float(temp_range[0])
            upper = float(temp_range[1])
            if lower <= temp_val <= upper:
                formula = df.loc[index, 'Formula'].replace("^", "**").replace("Temp", "{Temp}")
                return formula
        else:
            if float(condition_temperature) == temp_val:
                formula = df.loc[index, 'Formula'].replace("^", "**").replace("Temp", "{Temp}")
                return formula
    return None


def find_formula_by_strain(df, strain_val):
    for index, row in df.iterrows():
        condition_strain = row['Strain']
        if '~' in condition_strain:
            strain_range = condition_strain.split('~')
            lower = float(strain_range[0])
            upper = float(strain_range[1])
            if lower <= strain_val <= upper:
                formula = df.loc[index, 'Formula'].replce("LOG", "*math.log10").replace("^", "**").replace("S", "{S}")
                return formula
        else:
            if float(condition_strain) == strain_val:
                formula = df.loc[index, 'Formula'].replce("LOG", "*math.log10").replace("^", "**").replace("S", "{S}")
                return formula
    return None


def concat_dataframes_with_gaps(df_lists, gap_rows=1):
    result_list = []
    empty_df = pd.DataFrame(np.nan, index=range(gap_rows), columns=df_lists[0].columns)
    for i, df in enumerate(df_lists):
        result_list.append(df)

        # 빈 행 추가 (마지막 그룹 제외)
        if i != len(df_lists)-1:
            result_list.append(empty_df)

    return pd.concat(result_list, axis=0).reset_index(drop=True)


def check_keyword_layout(x):
    if isinstance(x, str):
        return KEYWORD[0] in x or KEYWORD[1] in x
    return False


def combine_column_values(df):
    last_index = df.index[-1] + 1  # xlsx에서는 1부터 시작

    # 각 열의 유효한 값들을 하나의 문자열로 결합
    combined_columns = {}

    for column in df.columns:
        # 각 열의 유효한 값만 필터링하고 문자열로 변환
        valid_values = [str(val) for val in df[column] if pd.notna(val) and str(val).strip() != '']
        result = []  # 중복 제거된 값들이 들어갈 리스트
        for value in valid_values:
            if value not in result:
                if 'Compressor' not in value:  # Compressor는 풀네임에서 제외
                    result.append(value)
        if result:  # result값이 유효하다면 값들을 '_'로 묶음
            full_name = '_'.join(result).upper().replace('-', '_').replace(' ', '_')  # Ansys에서는 대문자 및 _로 표현됨
            combined_columns[(last_index, column+1)] = full_name  # xlsx에서는 1부터 시작
    return combined_columns


def split_dataframe(df, chunk_size=3):
    # 전체 행 수 확인
    total_rows = len(df)

    # 분리된 데이터프레임을 저장할 리스트
    df_list = []

    # chunk_size만큼 분리
    for i in range(0, total_rows, chunk_size):
        chunk = df.iloc[i:i + chunk_size]
        df_list.append(chunk)

    return df_list


def read_excel_with_merged_cells(ws):
    # 병합된 셀 정보 가져오기
    merged_ranges = ws.merged_cells.ranges

    # 병합된 셀의 값을 저장할 딕셔너리
    merged_values = {}

    # 병합된 셀의 값을 처리
    for merged_range in merged_ranges:
        # 병합 범위의 첫 번째 셀 값 가져오기
        first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        value = first_cell.value

        # 병합된 모든 셀에 같은 값 적용
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value

    # 데이터프레임으로 변환할 데이터 준비
    data = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            # 병합된 셀이면 저장된 값 사용, 아니면 셀 값 사용
            if (cell.row, cell.column) in merged_values:
                row_data.append(merged_values[(cell.row, cell.column)])
            else:
                row_data.append(cell.value)
        data.append(row_data)

    # 데이터프레임 생성
    df_data = pd.DataFrame(data)

    # KTP와 2D_Layout 행만 추출
    mask = df_data.iloc[:, 0].apply(check_keyword_layout)
    filtered_df = df_data[mask]
    named_dfs = split_dataframe(filtered_df, chunk_size=3)
    dict_list = [combine_column_values(named_df) for named_df in named_dfs]
    merged_dict = {}
    for d in dict_list:
        merged_dict.update(d)
    return merged_dict


def find_changed_points(data: np.array):
    diff = np.diff(data)
    last_index = data.size - 1
    sign, changed_points = None, [0]  # 첫번째 포인트
    for index, diff in enumerate(diff):
        if diff != 0:
            current_sign = 'positive' if diff > 0 else 'negative'
            if sign != current_sign:
                if sign is not None:
                    changed_points.append(index)
            sign = current_sign
    changed_points.append(last_index)
    return changed_points


def rainflow_cycle(df):
    # 타입 변환을 한 번만 수행하고 재사용
    df_max = df.iloc[1:].reset_index(drop=True)
    df_max['Stress'] = df_max['Stress'].astype(float)  # 한 번만 형변환

    # 변경점 찾기
    changed_indices = find_changed_points(df_max['Stress'])

    # 효율적인 리스트 생성
    reversal_values = df_max.loc[changed_indices, 'Stress'].tolist()

    # rainflow_counting 함수 호출 시 불필요한 리스트 컴프리헨션 제거
    df_cycle = rainflow_counting(reversal_values)
    return df_cycle, df_max, changed_indices


def rainflow_point(result_path, part_name, df, changed_indices):
    # 그래프 생성
    fig, ax = plt.subplots(figsize=(10, 6))

    # 데이터프레임 열 그래프로 표시
    ax.plot(df['Time'], df['Stress'], color='black', linewidth=1)

    x_changed = df.loc[changed_indices, 'Time'].values.tolist()
    y_changed = df.loc[changed_indices, 'Stress'].values.tolist()
    # 지정된 인덱스에 X 마커 표시
    # ax.scatter(df.loc[changed_indices, 'No'], df.loc[changed_indices, 'Stress'], marker='o', color='red', s=10, linewidth=2, label='changed index')
    ax.plot(x_changed, y_changed, '--o', color='red', ms=3, linewidth=1, label='changed index')

    # 그래프 스타일 및 레이블 설정
    ax.set_title(f"Rainflow cycle of {part_name}", fontsize=14)
    ax.set_xlabel('Step', fontsize=12)
    ax.set_ylabel('Stress[MPa]', fontsize=12)
    ax.grid(True, linestyle='--', alpha=0.7)
    ax.legend(loc='upper right')
    plt.tight_layout()

    plt.savefig(f'{result_path}/{part_name}.png', format='png')  # result 폴더가 있어야 함
    plt.cla()  # clear the current axes
    plt.clf()  # clear the current figure
    plt.close()  # closes the current figure


# def rainflow_counting(data):
#     """
#     Rainflow counting 알고리즘을 구현한 함수
#
#     Args:
#         data: 시계열 데이터 리스트
#
#     Returns:
#         ranges: 각 cycle의 range (최대값 - 최소값)
#         means: 각 cycle의 평균값
#         cycles: 각 cycle의 발생 횟수
#         max_vals: 각 cycle의 최대값
#         min_vals: 각 cycle의 최소값
#     """
#     # scipy의 find_peaks 함수를 사용하여 피크(극댓값)와 골(극솟값) 찾기
#     indices = signal.find_peaks_cwt(np.abs(np.diff(data)), np.arange(1, 2))
#     indices = np.append(indices, len(data) - 1)  # 마지막 포인트 추가
#     indices = np.insert(indices, 0, 0)  # 첫 번째 포인트 추가
#
#     # 극값만 추출
#     extrema = [data[i] for i in indices]
#
#     # Rainflow counting
#     lst_data = []
#
#     # 스택 사용
#     stack = []
#
#     for i, point in enumerate(extrema):
#         if len(stack) < 2:
#             stack.append(point)
#             continue
#
#         # 스택의 마지막 두 요소
#         a, b = stack[-2], stack[-1]
#
#         # Range 계산
#         X = abs(b - a)
#         Y = abs(point - b)
#
#         if X >= Y:  # cycle 완성
#             # Range와 Mean 계산
#             cycle_range = X
#             cycle_mean = (a + b) / 2
#
#             # 최대값, 최소값 계산
#             cycle_max = max(a, b)
#             cycle_min = min(a, b)
#
#             lst_data.append([1.0, cycle_max, cycle_min, cycle_mean, cycle_range])  # 완전한 cycle은 1.0
#
#             # 스택에서 제거
#             stack.pop()
#             stack.pop()
#
#             # 새로운 포인트와 이전 포인트들 비교하기 위해 다시 스택에 추가
#             if len(stack) > 0:
#                 stack.append(point)
#             else:
#                 stack = [a, point]
#         else:
#             stack.append(point)
#
#     # 스택에 남은 요소들 처리 (half cycle)
#     for i in range(0, len(stack) - 1):
#         a, b = stack[i], stack[i + 1]
#         cycle_range = abs(b - a)
#         cycle_mean = (a + b) / 2
#         cycle_max = max(a, b)
#         cycle_min = min(a, b)
#
#         lst_data.append([0.5, cycle_max, cycle_min, cycle_mean, cycle_range])  # 반 cycle은 0.5
#
#     return pd.DataFrame(lst_data, columns=['Cycle', 'Max', 'Min', 'Mean', 'Range'])

def rainflow_counting(data):
    """
    Rainflow counting 알고리즘을 구현한 함수 - 첫 번째 점과 끝점을 cycle에 포함

    Args:
        data: 시계열 데이터 리스트

    Returns:
        DataFrame: Cycle, Max, Min, Mean, Range 컬럼을 포함하는 데이터프레임
    """
    # scipy의 find_peaks 함수를 사용하여 피크(극댓값)와 골(극솟값) 찾기
    indices = signal.find_peaks_cwt(np.abs(np.diff(data)), np.arange(1, 2))
    indices = np.append(indices, len(data) - 1)  # 마지막 포인트 추가
    indices = np.insert(indices, 0, 0)  # 첫 번째 포인트 추가

    # 극값만 추출
    extrema = [data[i] for i in indices]

    # Rainflow counting
    lst_data = []

    # 스택 사용
    stack = []

    for i, point in enumerate(extrema):
        if len(stack) < 2:
            stack.append(point)
            continue

        # 스택의 마지막 두 요소
        a, b = stack[-2], stack[-1]

        # Range 계산
        X = abs(b - a)
        Y = abs(point - b)

        if X >= Y:  # cycle 완성
            # Range와 Mean 계산
            cycle_range = X
            cycle_mean = (a + b) / 2

            # 최대값, 최소값 계산
            cycle_max = max(a, b)
            cycle_min = min(a, b)

            lst_data.append([1.0, cycle_max, cycle_min, cycle_mean, cycle_range])  # 완전한 cycle은 1.0

            # 스택에서 제거
            stack.pop()
            stack.pop()

            # 새로운 포인트와 이전 포인트들 비교하기 위해 다시 스택에 추가
            if len(stack) > 0:
                stack.append(point)
            else:
                stack = [a, point]
        else:
            stack.append(point)

    # 스택에 남은 요소들 처리 (half cycle)
    while len(stack) >= 2:
        # 스택에서 첫 번째와 두 번째 요소 가져오기
        a, b = stack[0], stack[1]

        # Range와 Mean 계산
        cycle_range = abs(b - a)
        cycle_mean = (a + b) / 2

        # 최대값, 최소값 계산
        cycle_max = max(a, b)
        cycle_min = min(a, b)

        lst_data.append([0.5, cycle_max, cycle_min, cycle_mean, cycle_range])  # 반 cycle은 0.5

        # 처리한 두 요소 제거
        stack.pop(0)

    # 첫 번째 점과 마지막 점 사이의 사이클 추가
    if len(extrema) >= 2:
        first_point = extrema[0]
        last_point = extrema[-1]

        cycle_range = abs(last_point - first_point)
        cycle_mean = (first_point + last_point) / 2
        cycle_max = max(first_point, last_point)
        cycle_min = min(first_point, last_point)

        # 시작점과 끝점 사이의 사이클 추가 (0.5 cycle로 처리)
        lst_data.append([0.5, cycle_max, cycle_min, cycle_mean, cycle_range])

    return pd.DataFrame(lst_data, columns=['Cycle', 'Max', 'Min', 'Mean', 'Range'])

def get_max_dataframe(df, col):
    # 1. 컬럼에서 최대값을 가진 행 찾기
    max_index = df[col].idxmax()
    # 2. 최대 Stress 값을 가진 행의 nodeID 값 추출
    target_node_id = df.loc[max_index]['NodeID']

    if col == "Temperature":
        column = COLUMNS_BASIC
    elif col == "Equivalent_Stress":
        column = ["Time", "NodeID", col, "Temperature", "Principal_1"]
    else:
        column = ["Time", "NodeID", col, "Temperature"]
    return df[df['NodeID'] == target_node_id].reset_index(drop=True)[column]


def extract_all_stress_temperature(df):
    df_stress = df.loc[df.groupby('Time')['Equivalent_Stress'].idxmax()].reset_index(drop=True)[COLUMNS_BASIC]
    df_temperature = df.loc[df.groupby('Time')['Temperature'].idxmax()].reset_index(drop=True)[COLUMNS_BASIC]
    lst_max_dataframes = [df_stress, df_temperature] + [get_max_dataframe(df=df, col=col_name) for col_name in COLUMNS_ORDER]
    max_stress = lst_max_dataframes[2].copy()
    sign_values = np.sign(max_stress["Principal_1"])
    max_stress['Equivalent_Stress'] = np.where(sign_values == 0, 1, sign_values) * max_stress['Equivalent_Stress']
    lst_max_dataframes.append(max_stress[COLUMNS_BASIC])
    lst_max_dataframes[2] = lst_max_dataframes[2][COLUMNS_BASIC]
    # df_stress, df_temperature, df_max_stress, df_max_temperature # max_principal # min_principal

    for df_dpf in lst_max_dataframes:
        df_dpf.columns.values[2] = 'Stress'

    return lst_max_dataframes


def extract_yellow_cells_info(excel_file):
    """
    엑셀 파일에서 노란색 셀이 있는 행의 첫 번째 값(또는 None인 경우 두 번째 값)과
    노란색 셀의 위치를 추출하는 함수

    Args:
        excel_file: 엑셀 파일 경로

    Returns:
        dict: 시트별로 노란색 셀 정보를 담은 딕셔너리
    """
    # 엑셀 파일 로드
    wb = openpyxl.load_workbook(excel_file)

    # 결과를 저장할 딕셔너리
    results = {}

    # 노란색 셀의 색상 코드 (FFFFFFCC)
    yellow_color = CELL_COLOR

    # 모든 시트에 대해 반복
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_results = []

        # 처리한 행을 추적하는 집합
        processed_rows = set()

        # 모든 셀 검사
        for row_idx, row in enumerate(ws.rows, 1):
            yellow_cells_in_row = []

            for col_idx, cell in enumerate(row, 1):
                # 셀이 노란색인지 확인
                if cell.fill.start_color.index == yellow_color:
                    # 셀 위치 (A1 형식)
                    cell_position = (row_idx, col_idx)
                    yellow_cells_in_row.append(cell_position)

            # 현재 행에 노란색 셀이 있고, 아직 처리되지 않은 경우
            if yellow_cells_in_row and row_idx not in processed_rows:
                # 해당 행의 첫 번째 셀 값
                first_cell_value = ws.cell(row=row_idx, column=1).value

                # 첫 번째 값이 None이면 두 번째 값 사용
                if first_cell_value is None:
                    first_cell_value = ws.cell(row=row_idx, column=2).value

                # 결과 딕셔너리에 추가
                row_info = {
                    "key": first_cell_value,
                    "yellow": yellow_cells_in_row
                }

                sheet_results.append(row_info)
                processed_rows.add(row_idx)

        results[sheet_name] = sheet_results

    return results


def reorder_data(node_ids, node_ids_2, data):
    """
    node_ids 순서에 맞게 data 배열을 재정렬합니다.

    Parameters:
    -----------
    node_ids : list
        stress의 인덱스 순서
    node_ids_2 : list
        temperature의 인덱스 순서
    temperature : list
        원본 temperature 값들

    Returns:
    --------
    list
        node_ids 순서에 맞게 재정렬된 temperature 배열
    """
    # node_ids_2의 각 요소가 어느 위치에 있는지 매핑
    node_ids_2_position = {node_id: idx for idx, node_id in enumerate(node_ids_2)}

    # node_ids의 순서에 맞게 temperature 재정렬
    reordered_data = []
    for node_id in node_ids:
        position = node_ids_2_position[node_id]
        reordered_data.append(data[position])
    return np.array(reordered_data)
