from ansys.dpf import core as dpf
import os
import openpyxl
import configparser
import time
import pandas as pd
from commons import isdir_and_make, merge_repeated_headers, process_pairs_and_differences, plot_columns_in_groups, read_excel_with_merged_cells, find_string_changes


properties = configparser.ConfigParser()  # 클래스 객체 생성


class DPF_Displacement:
    def __init__(self, result_path, input_file, rst_path, save_data, save_figure):
        isdir_and_make(result_path)  # result 폴더 생성

        self.model = None
        self.mesh = None
        self.result_path = result_path
        self.input_file = input_file
        self.rst_path = rst_path
        self.save_data = True if save_data == 'Yes' else False
        self.save_figure = True if save_figure == 'Yes' else False

        self.exe_time = None
        self.sheet_name = []
        self.points = {}
        self.nodes = {}
        self.time_step = {}
        self.time_data = None
        self.time_sec = []
        self.step_ids = []
        self.coordinates = {}

    def run(self):
        print("************************************************************")
        print("*** DPF Displacement Analysis Start!\n"
              "*** Please Do not try additional command until it completes")
        print("************************************************************\n")
        print(f"[INFO] Save CSV data option: {self.save_data}")
        print(f"[INFO] Save Figure option: {self.save_figure}")
        print(f"[INFO] Result Location: {self.result_path}")
        print(f"[INFO] RST folder: {self.rst_path}\n")
        self.exe_time = time.strftime('%Y%m%d_%H%M%S', time.localtime(time.time()))
        print(f'Starting at: {self.exe_time}')

        if self.save_data is True or self.save_figure is True:
            isdir_and_make(f'{self.result_path}/{self.exe_time}')  # result 결과 시간 폴더 생성

        self._load_in_point_node()
        self._load_rst()
        self._get_time_frequencies()
        dict_dis = self._extract_step_displacement()

        # 1. 파일 생성
        writer = pd.ExcelWriter(f"{self.result_path}/result_disp_{self.exe_time}.xlsx", engine='openpyxl')
        for sn in self.sheet_name:
            if 'Time' not in sn:
                df_coord = self._get_dataframe_coordinates(sn)
                df_coord.to_excel(writer, sheet_name=f'coordinate_{sn}', index=False, header=False)

        displacement_axis = ['X', 'Y']
        diff_sheet = []
        for sn in self.sheet_name:
            if 'Time' not in sn:
                for axis in displacement_axis:
                    df_disp, diff_state = self.get_dataframe_dis(dict_dis, axis, sn)
                    disp_sheet_name = f"disp_{axis}_{sn}"
                    df_disp.to_excel(writer, sheet_name=disp_sheet_name, index=False, header=False)
                    if diff_state is True:
                        diff_sheet.append(disp_sheet_name)

        # 3. 작성 완료 후 파일 저장
        writer.close()

        # 4. 일부 쉬트 merge 셀 적용
        if diff_sheet:
            merge_repeated_headers(f"{self.result_path}/result_disp_{self.exe_time}.xlsx", diff_sheet)

        self.stop()

    def stop(self):
        print(f'Ending at: {time.strftime("%a, %d-%b-%Y %I:%M:%S", time.localtime(time.time()))}')
        print("************************************************************")
        print("*** DPF Displacement Analysis completed")
        print("************************************************************\n")

    def get_dataframe_dis(self, dict_dis: dict, axis: str, sn: str):
        time_data = dict_dis.keys()
        disp = [[df.loc[node, axis] for node in self.nodes[sn]] for df in dict_dis.values()]
        if '-A' in self.points[sn][0] or '-B' in self.points[sn][0]:
            diff_apply = True
            node_by_point = dict(zip(self.points[sn], self.nodes[sn]))
            organized, organized_diff, organized_nodes, point_main, point_sub, units = [], [], [], [], [], []
            for point in self.points[sn]:
                if '-A' in point:
                    B_point = point.replace('-A', '-B')
                    organized.append(point)
                    organized.append(B_point)

                    organized_diff.append(point)
                    organized_diff.append(B_point)
                    organized_diff.append(point.replace('-A', 'A-B'))

                    organized_nodes.append(node_by_point[point])
                    organized_nodes.append(node_by_point[B_point])
                    organized_nodes.append('')

                    point_sub.append('A')
                    point_sub.append('B')
                    point_sub.append('A-B')

                    units.append(f'Disp[{self.mesh.unit}]')
                    units.append(f'Disp[{self.mesh.unit}]')
                    units.append(f'Diff.[{self.mesh.unit}]')

                    name = point.replace('-A', '')
                    for _ in range(3):
                        point_main.append(name)

            lst_disp_diff = [process_pairs_and_differences(values) for values in pd.DataFrame(data=disp, columns=self.points[sn]).reindex(columns=organized).values]
            # points간 diff 포함 그래프 그리기
            if self.save_figure is True:
                plot_columns_in_groups(df=pd.DataFrame(data=lst_disp_diff, columns=organized_diff, index=list(time_data)),
                                       filepath=f'{self.result_path}/{self.exe_time}',
                                       dict_changes=self.time_step,
                                       x_data=self.time_data)

            point_col1 = ['Points'] + point_main
            point_col2 = ['Points'] + point_sub
            node_col = ['Node'] + organized_nodes
            name_col = ['Time'] + units

            disp_rev = [point_col1, point_col2, node_col, name_col]
            for t, dis in zip(time_data, lst_disp_diff):
                disp_rev.append([t] + dis)
        else:
            diff_apply = False
            point_col = ['Points'] + self.points[sn]
            node_col = ['Node'] + self.nodes[sn]
            name_col = ['Time'] + [f'Disp[{self.mesh.unit}]' for _ in range(len(self.nodes[sn]))]

            disp_rev = [point_col, node_col, name_col]
            for t, dis in zip(time_data, disp):
                disp_rev.append([t] + dis)

        return pd.DataFrame(disp_rev), diff_apply

    def _load_in_point_node(self):
        wb = openpyxl.load_workbook(self.input_file)
        self.sheet_name = wb.sheetnames  # the list that stores the sheetnames
        for sn in self.sheet_name:
            if 'Time' in sn:
                df_time = read_excel_with_merged_cells(wb, sn)
                self.time_data = df_time.iloc[:, 0]
                self.time_step = find_string_changes(df_time, df_time.columns[1])
            else:
                df_input = pd.read_excel(self.input_file, engine='openpyxl', sheet_name=sn)
                self.points[sn] = list(df_input.iloc[:, 0])
                self.nodes[sn] = list(df_input.iloc[:, 1])

    def _load_rst(self):
        # rst 파일 찾기
        rst_file = [file for file in os.listdir(self.rst_path) if '.rst' in file]
        # rst 파일 직접 로드
        ds = dpf.DataSources(os.path.join(self.rst_path, rst_file[0]))
        self.model = dpf.Model(ds)
        print(self.model)

    def _get_time_frequencies(self):
        metadata = self.model.metadata
        self.mesh = metadata.meshed_region

        self.coordinates = {node.id: node.coordinates for node in self.mesh.nodes}

        # time_frequencies 객체 가져오기
        time_freq = metadata.time_freq_support.time_frequencies
        self.time_sec = time_freq.data
        self.step_ids = time_freq.scoping.ids

    def _get_dataframe_coordinates(self, sn):
        data_coord = [['Points', 'Node', 'Location X', 'Location Y']]
        for point, node in zip(self.points[sn], self.nodes[sn]):
            coordinates = self.coordinates[node]
            data_coord.append([point, node, coordinates[0], coordinates[1]])
        return pd.DataFrame(data_coord)

    def _extract_step_displacement(self):
        disp_op = self.model.results.displacement()
        time_index = 0
        dict_dis = {}
        for step in self.step_ids:
            time_steps_scoping = dpf.time_freq_scoping_factory.scoping_by_load_step(step)
            # disp = model.results.displacement.on_time_scoping(time_steps_scoping).eval()
            disp_op.inputs.time_scoping(time_steps_scoping)
            disp = disp_op.outputs.fields_container()
            for dis in disp:
                if float(self.time_sec[time_index]).is_integer():
                    node_ids = dis.scoping.ids
                    df = pd.DataFrame(data=dis.data, index=node_ids, columns=['X', 'Y', 'Z'])
                    df_dis = df.sort_index(ascending=True)
                    if self.save_data is True:
                        df_dis.to_csv(f'{self.result_path}/{self.exe_time}/displacement_step{step}_{round(self.time_sec[time_index], 1)}s.csv')
                    dict_dis[int(self.time_sec[time_index])] = df_dis
                time_index += 1
        return dict_dis


if __name__ == "__main__":
    properties.read('config.ini', encoding="UTF-8")  # 파일 읽기
    fm_params = properties["ENV"]  # 섹션 선택

    dpf_analysis = DPF_Displacement(result_path=fm_params['result_path'].strip(),
                                    input_file=fm_params['input_file_path'].strip(),
                                    rst_path=fm_params['rst_folder_path'].strip(),
                                    save_data=fm_params['save_raw_data'].strip(),
                                    save_figure=fm_params['save_figure'].strip())
    dpf_analysis.run()
