import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Alignment


COLOR_LIST = ['#808080',  # Gray
              '#3F51B5',  # Indigo
              '#03A9F4',  # Light Blue
              '#4CAF50',  # Green
              '#FFEB3B',  # Yellow
              '#FF00FF',  # Magenta
              '#F44336',  # Red
              '#00BCD4',  # Cyan
              '#FF9800'   # Orange
              '#E91E63',  # Pink
              '#2196F3',  # Blue
              '#9C27B0',  # Purple
              '#673AB7',  # Deep Purple
              '#009688',  # Teal
              '#8BC34A',  # Light Green
              '#CDDC39',  # Lime
              '#FFC107',  # Amber
              '#FF5722'  # Deep Orange
              ]

ALPHA_AREA = 0.1
ALPHA_LINE = 0.3


def isdir_and_make(dir_name: str):
    if not (os.path.isdir(dir_name)):
        os.makedirs(name=dir_name, exist_ok=True)
        print(f"Success: Create {dir_name}\n")


def remove_duplicates(lst):
    result = []
    for item in lst:
        if item not in result:
            result.append(item)
    return result


def process_pairs_and_differences(arr: np.array):
    # 원본 배열의 길이가 짝수인지 확인
    if len(arr) % 2 != 0:
        # 홀수인 경우, 마지막 요소를 제외
        arr = arr[:-1]

    # 1. 두 개씩 데이터 추출 (2열로 재구성)
    pairs = arr.reshape(-1, 2)

    # 2. 각 쌍의 차이 계산 A-B
    differences = pairs[:, 0] - pairs[:, 1]

    # 3. 원본 쌍과 차이를 합치기
    result = list(np.column_stack((pairs, differences)).flatten())

    return result


def plot_columns_in_groups(df, filepath,  dict_changes, x_data, columns_per_plot=3):
    """
    DataFrame의 컬럼들을 지정된 수만큼 그룹으로 나누어 그래프를 그립니다.

    Parameters:
    -----------
    df : pandas DataFrame
        그래프를 그릴 데이터프레임
    columns_per_plot : int
        각 그래프당 표시할 컬럼 수
    figsize : tuple
        각 그래프의 크기 (가로, 세로)
    """

    # 모든 컬럼 가져오기
    all_columns = df.columns.tolist()

    # 컬럼들을 지정된 수만큼의 그룹으로 나누기
    number_of_groups = len(all_columns) // columns_per_plot

    # 색상 설정
    colors = ['red', 'blue', 'green']  # A, B, A-B

    # 각 그룹별로 그래프 그리기
    for i in range(number_of_groups):
        start_idx = i * columns_per_plot
        end_idx = start_idx + columns_per_plot
        current_columns = all_columns[start_idx:end_idx]
        title_name = current_columns[0].replace('-A', '')
        fig, ax = plt.subplots(figsize=(12, 6))

        # 각 컬럼에 대해 라인 플롯 생성
        for j, column in enumerate(current_columns):
            ax.plot(df.index, df[column],
                    label=column,
                    color=colors[j],
                    linewidth=2,
                    marker='o',
                    markersize=4)

        # 그래프 꾸미기
        ax.set_title(f"{title_name} result", fontsize=18, pad=15)
        ax.set_xlabel('Time(sec)', fontsize=12)
        ax.set_ylabel('Displacement(mm)', fontsize=12)
        ax.set_xlim(0, df.index.max())  # x축 범위를 데이터의 최소값과 최대값으로 설정 x_data.min(), x_data.max()
        ax.legend(fontsize=10, bbox_to_anchor=(1.01, 1), loc='upper left')
        plt.grid(True, linestyle='--', alpha=0.7)

        if dict_changes != {}:  # 비어있는지 확인
            # 영역 구분
            locations = []
            ratings = list(dict_changes.values())
            color_rate = {rate: color for rate, color in zip(remove_duplicates(ratings), COLOR_LIST)}
            for indices, rating in dict_changes.items():
                ax.axvspan(x_data[indices[0]], x_data[indices[1]], alpha=ALPHA_AREA, color=color_rate[rating])
                if indices[0] != 0:
                    ax.axvline(x=x_data[indices[0]], color='black', linestyle='--', linewidth=1, alpha=ALPHA_LINE)
                locations.append(np.mean([x_data[indices[0]], x_data[indices[1]]]))

            ax_twin = ax.twiny()
            ax_twin.set_xlim(ax.get_xlim())
            ax_twin.set_xticks(np.array(locations))
            ax_twin.tick_params(axis='x', length=0)  # x축의 tick length를 0으로 설정
            ax_twin.set_xticklabels(ratings)

        # 여백 조정
        plt.tight_layout()
        plt.savefig(f'{filepath}/{title_name}.png', format='png')
        # plt.savefig(f'{filepath}/{filename}.svg', format='svg')
        plt.cla()  # clear the current axes
        plt.clf()  # clear the current figure
        plt.close()  # closes the current figure


def merge_repeated_headers(excel_path: str, sheet_names: list):
    """
    엑셀 파일의 첫 번째 행에서 반복되는 값들을 병합합니다.

    Parameters:
    -----------
    excel_path : str
        입력 엑셀 파일 경로
    output_path : str
        저장할 엑셀 파일 경로
    """
    # 엑셀 파일을 openpyxl로 읽기
    wb = openpyxl.load_workbook(excel_path)

    for sh in sheet_names:
        ws = wb[sh]

        # 첫 번째 행의 값들 가져오기
        first_row = list(ws[1])

        # 병합할 셀 범위 찾기
        merge_ranges = []
        start_col = 0
        current_value = first_row[0].value

        for i in range(1, len(first_row)):
            if first_row[i].value != current_value:
                if i - start_col > 1:  # 2개 이상의 셀이 같은 값을 가질 때만 병합
                    merge_ranges.append((start_col + 1, i))
                start_col = i
                current_value = first_row[i].value

        # 마지막 범위 처리
        if len(first_row) - start_col > 1:
            merge_ranges.append((start_col + 1, len(first_row)))

        # 셀 병합 및 스타일 적용
        for start, end in merge_ranges:
            merge_range = f"{chr(64 + start)}1:{chr(64 + end)}1"
            ws.merge_cells(merge_range)
            merged_cell = ws[f"{chr(64 + start)}1"]
            merged_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A1:A2')  # 첫번째 열에서의 Point 병합
        # 병합된 셀에 정렬 설정
        merged_cell = ws['A1']
        merged_cell.alignment = Alignment(horizontal='center',
                                          vertical='center')
    # 결과 저장
    wb.save(excel_path)


def read_excel_with_merged_cells(wb, sheet_name):
    # 워크북 로드
    ws = wb[wb.sheetnames[sheet_name] if isinstance(sheet_name, int) else sheet_name]

    # 병합된 셀 정보 가져오기
    merged_ranges = ws.merged_cells.ranges

    # 병합된 셀의 값을 저장할 딕셔너리
    merged_values = {}

    # 병합된 셀의 값을 처리
    for merged_range in merged_ranges:
        # 병합 범위의 첫 번째 셀 값 가져오기
        first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        value = first_cell.value

        # 병합된 모든 셀에 같은 값 적용
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value

    # 데이터프레임으로 변환할 데이터 준비
    data = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            # 병합된 셀이면 저장된 값 사용, 아니면 셀 값 사용
            if (cell.row, cell.column) in merged_values:
                row_data.append(merged_values[(cell.row, cell.column)])
            else:
                row_data.append(cell.value)
        if all(x is None for x in row_data) is False:
            data.append(row_data)

    # 데이터프레임 생성
    df = pd.DataFrame(data)
    df.columns = df.iloc[0]
    df = df.iloc[1:]

    return df


def find_string_changes(df, column_name):
    """
    데이터프레임의 문자열 컬럼에서 값이 변하는 지점을 찾아
    변화가 발생한 문자열과 해당 인덱스를 반환합니다.

    Parameters:
    df (pandas.DataFrame): 분석할 데이터프레임
    column_name (str): 분석할 문자열 컬럼명

    Returns:
    tuple: (변화가 발생한 인덱스 리스트, 변화가 발생한 문자열 리스트)
    """
    # 이전 값과 다른 값을 가진 행 찾기
    value_changes = df[column_name] != df[column_name].shift()

    # 첫 번째 행은 항상 포함 (초기값)
    value_changes.iloc[0] = True

    # 변화가 발생한 지점의 인덱스와 문자열 추출
    change_indices = df.index[value_changes].tolist()
    change_indices.append(df.index[-1])
    change_values = df.loc[value_changes, column_name].tolist()

    return {(change_indices[i], change_indices[i+1]): change_values[i] for i in range(len(change_indices) - 1)}